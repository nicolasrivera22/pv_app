from __future__ import annotations

from io import BytesIO
from pathlib import Path
import shutil
import unicodedata
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo

from pv_product.utils import DEFAULT_CONFIG, build_7x24_from_excel, solar_profile_24

from .runtime_paths import bundled_workbook_path
from .types import LoadedConfigBundle
from .types import ValidationIssue
from .validation import validate_config

SHEET_ALIASES = {
    "Config": ("Config",),
    "Perfiles": ("Perfiles", "Profiles"),
    "Catalogos": ("Catalogos", "Catalogs"),
}

TABLE_COLUMNS = {
    "Demand_Profile": ["DOW", "HOUR", "RES", "IND", "TOTAL_kWh"],
    "Demand_Profile_General": ["HOUR", "RES", "IND", "TOTAL_kWh"],
    "Demand_Profile_Weights": ["HOUR", "W_RES", "W_IND", "W_RES_BASE", "W_IND_BASE", "W_TOTAL", "TOTAL_kWh"],
    "Month_Demand_Profile": ["MONTH", "Demand_month", "HSP_month"],
    "SUN_HSP_PROFILE": ["HOUR", "SOL"],
    "Precios_kWp_relativos": ["MIN", "MAX", "PRECIO_POR_KWP"],
    "Precios_kWp_relativos_Otros": ["MIN", "MAX", "PRECIO_POR_KWP"],
    "Inversor_Catalog": ["name", "AC_kW", "Vmppt_min", "Vmppt_max", "Vdc_max", "Imax_mppt", "n_mppt", "price_COP"],
    "Battery_Catalog": ["name", "nom_kWh", "max_kW", "max_ch_kW", "max_dis_kW", "price_COP"],
}
TABLE_FILE_MAP = {
    "Config": "Config.csv",
    "Demand_Profile": "Demand_Profile.csv",
    "Demand_Profile_General": "Demand_Profile_General.csv",
    "Demand_Profile_Weights": "Demand_Profile_Weights.csv",
    "Month_Demand_Profile": "Month_Demand_Profile.csv",
    "SUN_HSP_PROFILE": "SUN_HSP_PROFILE.csv",
    "Precios_kWp_relativos": "Precios_kWp_relativos.csv",
    "Precios_kWp_relativos_Otros": "Precios_kWp_relativos_Otros.csv",
    "Inversor_Catalog": "Inversor_Catalog.csv",
    "Battery_Catalog": "Battery_Catalog.csv",
}

CONFIG_KEY_ALIASES = {
    "coupling": "bat_coupling",
}

CONFIG_BOOL_FIELDS = {key for key, value in DEFAULT_CONFIG.items() if isinstance(value, bool)}
CONFIG_CHOICE_MAPS = {
    "pricing_mode": {"variable": "variable", "total": "total"},
    "kWp_seed_mode": {"auto": "auto", "manual": "manual"},
    "limit_peak_month_mode": {"max": "max", "maximo": "max", "fixed": "fixed", "fijo": "fixed"},
    "limit_peak_basis": {
        "weighted mean": "weighted_mean",
        "weighted_mean": "weighted_mean",
        "max": "max",
        "dia semana": "weekday",
        "weekday": "weekday",
        "p95": "p95",
    },
    "bat_coupling": {"ac": "ac", "dc": "dc"},
}
PROFILE_MODE_MAP = {
    "perfil horario relativo": "perfil horario relativo",
    "perfil hora dia de semana": "perfil hora dia de semana",
    "perfil general": "perfil general",
}


class WorkbookContractError(ValueError):
    """Raised when the uploaded workbook does not satisfy the expected contract."""


def _slug(value: Any) -> str:
    value = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
    return " ".join(value.strip().lower().replace("_", " ").split())


def _coerce_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(value)
    slug = _slug(value)
    if slug in {"true", "si", "yes", "y", "1"}:
        return True
    if slug in {"false", "no", "0"}:
        return False
    return None


def _normalize_mode(value: Any) -> str:
    if isinstance(value, bool):
        return "perfil hora dia de semana" if value else "perfil general"
    slug = _slug(value)
    if "relativo" in slug:
        return "perfil horario relativo"
    if "dia" in slug or "semana" in slug:
        return "perfil hora dia de semana"
    return "perfil general"


def _normalize_coupling(value: Any) -> str:
    slug = _slug(value)
    return "dc" if slug == "dc" else "ac"


def _normalize_choice(value: Any, mapping: dict[str, str], default: str) -> str:
    slug = _slug(value)
    return mapping.get(slug, default)


def _normalize_config_value(key: str, value: Any) -> Any:
    default_value = DEFAULT_CONFIG.get(key)
    if key == "use_excel_profile":
        if isinstance(value, bool):
            return _normalize_mode(value), None
        normalized = _normalize_mode(value)
        if normalized not in PROFILE_MODE_MAP.values():
            normalized = "perfil general"
        if _slug(value) in {"", "perfil horario relativo", "perfil hora dia de semana", "perfil general"} or normalized != "perfil general" or _slug(value) == "perfil general":
            return normalized, None
        return normalized, f"Valor inválido para '{key}': {value!r}."

    if key in CONFIG_CHOICE_MAPS:
        slug = _slug(value)
        if slug in CONFIG_CHOICE_MAPS[key]:
            return CONFIG_CHOICE_MAPS[key][slug], None
        return default_value, f"Valor inválido para '{key}': {value!r}."

    if key in CONFIG_BOOL_FIELDS:
        bool_value = _coerce_bool(value)
        if bool_value is not None:
            return bool_value, None
        return default_value, f"Valor booleano inválido para '{key}': {value!r}."

    if isinstance(default_value, int) and not isinstance(default_value, bool):
        return int(float(value)), None
    if isinstance(default_value, float):
        return float(value), None
    return value, None


def _open_sources(path_or_bytes: str | bytes | Path | BytesIO) -> tuple[Any, Any, str]:
    if isinstance(path_or_bytes, BytesIO):
        data = path_or_bytes.getvalue()
        return BytesIO(data), BytesIO(data), "uploaded.xlsx"
    if isinstance(path_or_bytes, bytes):
        return BytesIO(path_or_bytes), BytesIO(path_or_bytes), "uploaded.xlsx"
    path = Path(path_or_bytes)
    return path, path, path.name


def _resolve_sheet_name(workbook, canonical_name: str) -> str:
    aliases = SHEET_ALIASES[canonical_name]
    for sheet_name in workbook.sheetnames:
        if sheet_name in aliases:
            return sheet_name
    raise WorkbookContractError(
        f"Falta la hoja '{canonical_name}'. Hojas aceptadas para este bloque: {', '.join(aliases)}."
    )


def read_table_from_excel(path_or_bytes: str | bytes | Path | BytesIO, sheet_name: str, table_name: str) -> pd.DataFrame:
    excel_source, workbook_source, _ = _open_sources(path_or_bytes)
    xls = pd.ExcelFile(excel_source)
    workbook = load_workbook(workbook_source, data_only=True)
    resolved_sheet_name = _resolve_sheet_name(workbook, sheet_name)
    ws = workbook[resolved_sheet_name]
    if table_name not in ws.tables:
        raise WorkbookContractError(
            f"Falta la tabla '{table_name}' en la hoja '{resolved_sheet_name}'."
        )
    table_obj = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table_obj.ref)
    usecols = f"{get_column_letter(min_col)}:{get_column_letter(max_col)}"
    frame = pd.read_excel(
        xls,
        sheet_name=resolved_sheet_name,
        header=min_row - 1,
        usecols=usecols,
        nrows=max_row - min_row + 1,
    )
    renamed = {
        col: str(col).split(".")[0]
        for col in frame.columns
        if str(col).rsplit(".", 1)[-1].isdigit()
    }
    frame = frame.rename(columns=renamed)
    frame = frame.replace(r"^\s*$", np.nan, regex=True).dropna(how="all").reset_index(drop=True)
    required_columns = TABLE_COLUMNS.get(table_name, [])
    missing_columns = [column for column in required_columns if column not in frame.columns]
    if missing_columns:
        raise WorkbookContractError(
            f"La tabla '{table_name}' en la hoja '{resolved_sheet_name}' no tiene las columnas requeridas: {', '.join(missing_columns)}."
        )
    if required_columns:
        frame = frame.dropna(subset=[required_columns[0]]).reset_index(drop=True)
    return frame


def _normalize_config(config: dict[str, Any]) -> tuple[dict[str, Any], list[ValidationIssue]]:
    cfg = dict(DEFAULT_CONFIG)
    issues: list[ValidationIssue] = []
    for key, value in config.items():
        internal_key = CONFIG_KEY_ALIASES.get(key, key)
        if internal_key not in cfg:
            continue
        normalized_value, error = _normalize_config_value(internal_key, value)
        cfg[internal_key] = normalized_value
        if error is not None:
            issues.append(ValidationIssue("error", internal_key, error))

    if "bat_coupling" not in cfg and "coupling" in config:
        cfg["bat_coupling"] = _normalize_coupling(config["coupling"])
    return cfg, issues


def _default_catalogs() -> tuple[pd.DataFrame, pd.DataFrame]:
    inverter_catalog = pd.DataFrame(
        [
            {"name": "INV-5k", "AC_kW": 5.0, "Vmppt_min": 200, "Vmppt_max": 800, "Vdc_max": 1000, "Imax_mppt": 13, "n_mppt": 2, "price_COP": 5_500_000},
            {"name": "INV-8k", "AC_kW": 8.0, "Vmppt_min": 250, "Vmppt_max": 850, "Vdc_max": 1000, "Imax_mppt": 18, "n_mppt": 2, "price_COP": 7_800_000},
            {"name": "INV-10k", "AC_kW": 10.0, "Vmppt_min": 250, "Vmppt_max": 850, "Vdc_max": 1100, "Imax_mppt": 18, "n_mppt": 2, "price_COP": 9_300_000},
            {"name": "INV-12k", "AC_kW": 12.0, "Vmppt_min": 300, "Vmppt_max": 900, "Vdc_max": 1100, "Imax_mppt": 22, "n_mppt": 2, "price_COP": 10_700_000},
            {"name": "INV-15k", "AC_kW": 15.0, "Vmppt_min": 350, "Vmppt_max": 1000, "Vdc_max": 1100, "Imax_mppt": 26, "n_mppt": 2, "price_COP": 13_500_000},
        ]
    )
    battery_catalog = pd.DataFrame(
        [
            {"name": "BAT-5", "nom_kWh": 5.0, "max_kW": 3.0, "max_ch_kW": 3.0, "max_dis_kW": 3.0, "price_COP": 7_000_000},
            {"name": "BAT-10", "nom_kWh": 10.0, "max_kW": 5.0, "max_ch_kW": 5.0, "max_dis_kW": 5.0, "price_COP": 12_500_000},
            {"name": "BAT-15", "nom_kWh": 15.0, "max_kW": 7.5, "max_ch_kW": 7.5, "max_dis_kW": 7.5, "price_COP": 17_500_000},
            {"name": "BAT-20", "nom_kWh": 20.0, "max_kW": 10.0, "max_ch_kW": 10.0, "max_dis_kW": 10.0, "price_COP": 21_500_000},
        ]
    )
    return inverter_catalog, battery_catalog


def _default_tables() -> dict[str, pd.DataFrame]:
    rows = []
    for dow in range(1, 8):
        for hour in range(24):
            res = 0.8 + (0.3 if hour in (7, 8) else 0.0) + (0.7 if hour in (18, 19, 20) else 0.0)
            ind = 1.0 if (dow <= 5 and (8 <= hour < 18)) else 0.1
            rows.append({"DOW": dow, "HOUR": hour, "RES": res, "IND": ind, "TOTAL_kWh": res + ind})
    general_rows = []
    for hour in range(24):
        res = 0.8 + (0.3 if hour in (7, 8) else 0.0) + (0.7 if hour in (18, 19, 20) else 0.0)
        ind = 1.0 if 8 <= hour < 18 else 0.1
        general_rows.append({"HOUR": hour, "RES": res, "IND": ind, "TOTAL_kWh": res + ind})
    weights_rows = []
    total_sum = sum(row["TOTAL_kWh"] for row in general_rows)
    for row in general_rows:
        weights_rows.append(
            {
                "HOUR": row["HOUR"],
                "W_RES": row["RES"] / sum(r["RES"] for r in general_rows),
                "W_IND": row["IND"] / sum(r["IND"] for r in general_rows),
                "W_RES_BASE": row["RES"] / sum(r["RES"] for r in general_rows),
                "W_IND_BASE": row["IND"] / sum(r["IND"] for r in general_rows),
                "W_TOTAL": row["TOTAL_kWh"] / total_sum,
                "TOTAL_kWh": row["TOTAL_kWh"] / total_sum,
            }
        )
    month_profile = pd.DataFrame(
        {"MONTH": list(range(1, 13)), "Demand_month": [1.0] * 12, "HSP_month": [DEFAULT_CONFIG["HSP"]] * 12}
    )
    sun_profile = pd.DataFrame({"HOUR": list(range(24)), "SOL": solar_profile_24()})
    prices = pd.DataFrame(
        [
            {"MIN": 1, "MAX": 5, "PRECIO_POR_KWP": 5_500_000},
            {"MIN": 5, "MAX": 8, "PRECIO_POR_KWP": 5_000_000},
            {"MIN": 8, "MAX": 10, "PRECIO_POR_KWP": 4_500_000},
            {"MIN": 10, "MAX": 13, "PRECIO_POR_KWP": 4_500_000},
            {"MIN": 13, "MAX": 100, "PRECIO_POR_KWP": 3_500_000},
        ]
    )
    prices_others = pd.DataFrame(
        [
            {"MIN": 1, "MAX": 5, "PRECIO_POR_KWP": 1_000_000},
            {"MIN": 5, "MAX": 8, "PRECIO_POR_KWP": 900_000},
            {"MIN": 8, "MAX": 10, "PRECIO_POR_KWP": 800_000},
            {"MIN": 10, "MAX": 13, "PRECIO_POR_KWP": 700_000},
            {"MIN": 13, "MAX": 100, "PRECIO_POR_KWP": 600_000},
        ]
    )
    return {
        "Demand_Profile": pd.DataFrame(rows),
        "Demand_Profile_General": pd.DataFrame(general_rows),
        "Demand_Profile_Weights": pd.DataFrame(weights_rows),
        "Month_Demand_Profile": month_profile,
        "SUN_HSP_PROFILE": sun_profile,
        "Precios_kWp_relativos": prices,
        "Precios_kWp_relativos_Otros": prices_others,
    }


def _bundle_from_frames(
    config: dict[str, Any],
    config_table: pd.DataFrame,
    inverter_catalog: pd.DataFrame,
    battery_catalog: pd.DataFrame,
    demand_profile: pd.DataFrame,
    demand_profile_weights: pd.DataFrame,
    demand_profile_general: pd.DataFrame,
    month_profile: pd.DataFrame,
    sun_profile: pd.DataFrame | None,
    cop_kwp_table: pd.DataFrame,
    cop_kwp_table_others: pd.DataFrame,
    source_name: str,
) -> LoadedConfigBundle:
    cfg, normalization_issues = _normalize_config(config)

    if sun_profile is not None and {"HOUR", "SOL"}.issubset(sun_profile.columns):
        ordered_sun = sun_profile.sort_values("HOUR")
        solar = ordered_sun["SOL"].to_numpy(dtype=float)
        solar = solar / solar.sum() if solar.sum() > 0 else solar_profile_24()
    else:
        solar = solar_profile_24()

    if not {"MONTH", "Demand_month", "HSP_month"}.issubset(month_profile.columns):
        raise ValueError("La tabla Month_Demand_Profile debe incluir MONTH, Demand_month y HSP_month.")

    hsp_month = np.ones(12, dtype=float) * float(DEFAULT_CONFIG["HSP"])
    demand_month_factor = np.ones(12, dtype=float)
    for _, row in month_profile.iterrows():
        month_idx = int(row["MONTH"]) - 1
        if 0 <= month_idx < 12:
            hsp_month[month_idx] = float(row["HSP_month"])
            demand_month_factor[month_idx] = float(row["Demand_month"])

    profile_mode = cfg["use_excel_profile"]
    if profile_mode == "perfil hora dia de semana":
        frame = demand_profile.rename(columns={"TOTAL_kWh": "TOTAL"})
        dow24, day_w = build_7x24_from_excel(frame)
    elif profile_mode == "perfil horario relativo":
        frame = demand_profile_weights.rename(columns={"TOTAL_kWh": "TOTAL"}).copy()
        dow = pd.DataFrame({"DOW": range(1, 8)})
        frame = frame.merge(dow, how="cross").sort_values(by=["DOW", "HOUR"]).reset_index(drop=True)
        dow24, day_w = build_7x24_from_excel(frame[["DOW", "HOUR", "TOTAL"]], total=True)
    else:
        frame = demand_profile_general.rename(columns={"TOTAL_kWh": "TOTAL"}).copy()
        dow = pd.DataFrame({"DOW": range(1, 8)})
        frame = frame.merge(dow, how="cross").sort_values(by=["DOW", "HOUR"]).reset_index(drop=True)
        dow24, day_w = build_7x24_from_excel(frame[["DOW", "HOUR", "TOTAL"]], total=True)

    cfg["HSP"] = float(hsp_month.mean())

    bundle = LoadedConfigBundle(
        config=cfg,
        inverter_catalog=inverter_catalog.copy(),
        battery_catalog=battery_catalog.copy(),
        solar_profile=np.asarray(solar, dtype=float),
        hsp_month=np.asarray(hsp_month, dtype=float),
        demand_profile_7x24=np.asarray(dow24, dtype=float),
        day_weights=np.asarray(day_w, dtype=float),
        demand_month_factor=np.asarray(demand_month_factor, dtype=float),
        cop_kwp_table=cop_kwp_table.copy(),
        cop_kwp_table_others=cop_kwp_table_others.copy(),
        config_table=config_table.copy(),
        demand_profile_table=demand_profile.copy(),
        demand_profile_general_table=demand_profile_general.copy(),
        demand_profile_weights_table=demand_profile_weights.copy(),
        month_profile_table=month_profile.copy(),
        sun_profile_table=sun_profile.copy() if sun_profile is not None else pd.DataFrame(),
        source_name=source_name,
    )
    issues = [*normalization_issues, *validate_config(bundle)]
    return LoadedConfigBundle(**{**bundle.__dict__, "issues": tuple(issues)})


def load_config_from_excel(path_or_bytes: str | bytes | Path | BytesIO) -> LoadedConfigBundle:
    config_table = read_table_from_excel(path_or_bytes, "Config", "Config")
    month_profile = read_table_from_excel(path_or_bytes, "Perfiles", "Month_Demand_Profile")
    demand_profile = read_table_from_excel(path_or_bytes, "Perfiles", "Demand_Profile")
    demand_profile_weights = read_table_from_excel(path_or_bytes, "Perfiles", "Demand_Profile_Weights")
    demand_profile_general = read_table_from_excel(path_or_bytes, "Perfiles", "Demand_Profile_General")
    cop_kwp_table = read_table_from_excel(path_or_bytes, "Perfiles", "Precios_kWp_relativos")
    cop_kwp_table_others = read_table_from_excel(path_or_bytes, "Perfiles", "Precios_kWp_relativos_Otros")
    inverter_catalog = read_table_from_excel(path_or_bytes, "Catalogos", "Inversor_Catalog")
    battery_catalog = read_table_from_excel(path_or_bytes, "Catalogos", "Battery_Catalog")

    try:
        sun_profile = read_table_from_excel(path_or_bytes, "Perfiles", "SUN_HSP_PROFILE")
    except ValueError:
        sun_profile = None

    source_name = _open_sources(path_or_bytes)[2]
    config: dict[str, Any] = {}
    for _, row in config_table.iterrows():
        item = str(row.get("Item", "")).strip()
        if not item:
            continue
        config[item] = row.get("Valor")

    return _bundle_from_frames(
        config=config,
        config_table=config_table,
        inverter_catalog=inverter_catalog,
        battery_catalog=battery_catalog,
        demand_profile=demand_profile,
        demand_profile_weights=demand_profile_weights,
        demand_profile_general=demand_profile_general,
        month_profile=month_profile,
        sun_profile=sun_profile,
        cop_kwp_table=cop_kwp_table,
        cop_kwp_table_others=cop_kwp_table_others,
        source_name=source_name,
    )


def load_example_config() -> LoadedConfigBundle:
    example_path = bundled_workbook_path()
    if example_path.exists():
        return load_config_from_excel(example_path)

    inverter_catalog, battery_catalog = _default_catalogs()
    tables = _default_tables()
    config = dict(DEFAULT_CONFIG)
    config["use_excel_profile"] = "Perfil Horario Relativo"
    config["bat_coupling"] = "ac"
    return _bundle_from_frames(
        config=config,
        config_table=pd.DataFrame(
            [{"Grupo": "", "Descripción": "", "Item": key, "Valor": value, "Unidad": ""} for key, value in config.items()]
        ),
        inverter_catalog=inverter_catalog,
        battery_catalog=battery_catalog,
        demand_profile=tables["Demand_Profile"],
        demand_profile_weights=tables["Demand_Profile_Weights"],
        demand_profile_general=tables["Demand_Profile_General"],
        month_profile=tables["Month_Demand_Profile"],
        sun_profile=tables["SUN_HSP_PROFILE"],
        cop_kwp_table=tables["Precios_kWp_relativos"],
        cop_kwp_table_others=tables["Precios_kWp_relativos_Otros"],
        source_name="default-example",
    )


def rebuild_config_bundle(
    base_bundle: LoadedConfigBundle,
    *,
    config: dict[str, Any] | None = None,
    config_table: pd.DataFrame | None = None,
    inverter_catalog: pd.DataFrame | None = None,
    battery_catalog: pd.DataFrame | None = None,
    demand_profile: pd.DataFrame | None = None,
    demand_profile_weights: pd.DataFrame | None = None,
    demand_profile_general: pd.DataFrame | None = None,
    month_profile: pd.DataFrame | None = None,
    sun_profile: pd.DataFrame | None = None,
    cop_kwp_table: pd.DataFrame | None = None,
    cop_kwp_table_others: pd.DataFrame | None = None,
) -> LoadedConfigBundle:
    return _bundle_from_frames(
        config=config or dict(base_bundle.config),
        config_table=config_table if config_table is not None else base_bundle.config_table,
        inverter_catalog=inverter_catalog if inverter_catalog is not None else base_bundle.inverter_catalog,
        battery_catalog=battery_catalog if battery_catalog is not None else base_bundle.battery_catalog,
        demand_profile=demand_profile if demand_profile is not None else base_bundle.demand_profile_table,
        demand_profile_weights=demand_profile_weights if demand_profile_weights is not None else base_bundle.demand_profile_weights_table,
        demand_profile_general=demand_profile_general if demand_profile_general is not None else base_bundle.demand_profile_general_table,
        month_profile=month_profile if month_profile is not None else base_bundle.month_profile_table,
        sun_profile=sun_profile if sun_profile is not None else base_bundle.sun_profile_table,
        cop_kwp_table=cop_kwp_table if cop_kwp_table is not None else base_bundle.cop_kwp_table,
        cop_kwp_table_others=cop_kwp_table_others if cop_kwp_table_others is not None else base_bundle.cop_kwp_table_others,
        source_name=base_bundle.source_name,
    )


def load_bundle_from_tables(table_root: str | Path, *, source_name: str = "project") -> LoadedConfigBundle:
    root = Path(table_root)

    def _read_csv(table_name: str, *, optional: bool = False) -> pd.DataFrame | None:
        path = root / TABLE_FILE_MAP[table_name]
        if not path.exists():
            if optional:
                return None
            raise FileNotFoundError(f"Falta la tabla canónica '{path.name}' en '{root}'.")
        return pd.read_csv(path)

    config_table = _read_csv("Config")
    month_profile = _read_csv("Month_Demand_Profile")
    demand_profile = _read_csv("Demand_Profile")
    demand_profile_weights = _read_csv("Demand_Profile_Weights")
    demand_profile_general = _read_csv("Demand_Profile_General")
    cop_kwp_table = _read_csv("Precios_kWp_relativos")
    cop_kwp_table_others = _read_csv("Precios_kWp_relativos_Otros")
    inverter_catalog = _read_csv("Inversor_Catalog")
    battery_catalog = _read_csv("Battery_Catalog")
    sun_profile = _read_csv("SUN_HSP_PROFILE", optional=True)

    config: dict[str, Any] = {}
    assert config_table is not None
    for _, row in config_table.iterrows():
        item = str(row.get("Item", "")).strip()
        if not item:
            continue
        config[item] = row.get("Valor")

    return _bundle_from_frames(
        config=config,
        config_table=config_table,
        inverter_catalog=inverter_catalog,
        battery_catalog=battery_catalog,
        demand_profile=demand_profile,
        demand_profile_weights=demand_profile_weights,
        demand_profile_general=demand_profile_general,
        month_profile=month_profile,
        sun_profile=sun_profile,
        cop_kwp_table=cop_kwp_table,
        cop_kwp_table_others=cop_kwp_table_others,
        source_name=source_name,
    )


def _add_table(ws, data_frame: pd.DataFrame, table_name: str, start_row: int, start_col: int) -> None:
    for col_idx, column in enumerate(data_frame.columns, start=start_col):
        cell = ws.cell(row=start_row, column=col_idx, value=column)
        cell.font = Font(bold=True)
    for row_offset, (_, row) in enumerate(data_frame.iterrows(), start=1):
        for col_offset, column in enumerate(data_frame.columns):
            ws.cell(row=start_row + row_offset, column=start_col + col_offset, value=row[column])
    end_row = start_row + len(data_frame)
    end_col = start_col + len(data_frame.columns) - 1
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def _build_template_workbook(path: Path) -> None:
    wb = Workbook()
    ws_config = wb.active
    ws_config.title = "Config"
    ws_profiles = wb.create_sheet("Perfiles")
    ws_catalogs = wb.create_sheet("Catalogos")

    config_rows = []
    for key, value in DEFAULT_CONFIG.items():
        external_key = "coupling" if key == "bat_coupling" else key
        external_value = value
        if key == "use_excel_profile":
            external_value = "Perfil Horario Relativo"
        elif key == "bat_coupling":
            external_value = "AC"
        config_rows.append({"Grupo": "", "Descripción": "", "Item": external_key, "Valor": external_value, "Unidad": ""})
    _add_table(ws_config, pd.DataFrame(config_rows), "Config", start_row=1, start_col=2)

    tables = _default_tables()
    _add_table(ws_profiles, tables["Demand_Profile"], "Demand_Profile", start_row=3, start_col=1)
    _add_table(ws_profiles, tables["Demand_Profile_General"], "Demand_Profile_General", start_row=3, start_col=8)
    _add_table(ws_profiles, tables["Demand_Profile_Weights"], "Demand_Profile_Weights", start_row=3, start_col=13)
    _add_table(ws_profiles, tables["Month_Demand_Profile"], "Month_Demand_Profile", start_row=3, start_col=21)
    _add_table(ws_profiles, tables["SUN_HSP_PROFILE"], "SUN_HSP_PROFILE", start_row=3, start_col=25)
    _add_table(ws_profiles, tables["Precios_kWp_relativos"], "Precios_kWp_relativos", start_row=3, start_col=28)
    _add_table(ws_profiles, tables["Precios_kWp_relativos_Otros"], "Precios_kWp_relativos_Otros", start_row=3, start_col=32)

    inverter_catalog, battery_catalog = _default_catalogs()
    _add_table(ws_catalogs, inverter_catalog, "Inversor_Catalog", start_row=1, start_col=1)
    _add_table(ws_catalogs, battery_catalog, "Battery_Catalog", start_row=1, start_col=10)
    wb.save(path)


def ensure_template(path: str | Path) -> None:
    destination = Path(path)
    example_path = bundled_workbook_path()
    if example_path.exists() and example_path.resolve() != destination.resolve():
        shutil.copyfile(example_path, destination)
        return
    _build_template_workbook(destination)
