"""
Funciones utilitarias y helpers compartidos por la app FV.

Incluye:
- Lectura de perfiles y configuración desde Excel.
- Construcción de perfiles horarios/estacionales.
- Wrappers de despacho diario (grid/zero-export/isla).
- Gráficas de resultados.
"""

from __future__ import annotations

import math
import os
import random
import re
from typing import Dict, List, Optional, Tuple

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from .dispatch import dispatch_day
from .models import Battery, DispatchConfig, PVSystem
from .optimizer import Optimizer
from .simulator import Simulator


def safe_div(a, b, default=0.0):
    """Devuelve `a/b` evitando divisiones por cero (retorna `default`)."""
    return a / b if (b is not None and abs(b) > 1e-12) else default


def ann_to_month_rate(r_annual: float) -> float:
    """Convierte una tasa anual efectiva a tasa mensual equivalente."""
    return (1.0 + r_annual) ** (1.0 / 12.0) - 1.0


def read_table_from_excel(path: str, sheet_name: str, table_name: str) -> pd.DataFrame:
    """Lee un rango nombrado (tabla) desde Excel respetando encabezados y columnas."""
    xls = pd.ExcelFile(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    table_obj = ws.tables[table_name]
    cfg_tb_range = table_obj.ref

    min_col, min_row, max_col, max_row = range_boundaries(cfg_tb_range)

    df = pd.read_excel(
        xls,
        sheet_name=sheet_name,
        header=min_row - 1,
        usecols=f"{ws.cell(row=1, column=min_col).column_letter}:"
        f"{ws.cell(row=1, column=max_col).column_letter}",
        nrows=max_row - min_row,
    )
    cols = df.columns
    for col in cols:
        if re.search(r"\.\d+$", col):
            df.rename(columns={col: col.split(".")[0]}, inplace=True)
    return df


def solar_profile_24(shift=0, exponent=1.0):
    """Genera una curva solar sintética de 24h (suma 1) con shift y exponente opcional."""
    s = np.zeros(24, dtype=float)
    for h in range(6, 19):
        x = (h - 6) / 12.0
        s[h] = max(math.sin(math.pi * x), 0.0)
    if exponent != 1.0:
        s = np.power(s, exponent)
    if shift != 0:
        s = np.roll(s, shift)
    s_sum = s.sum()
    return s / s_sum if s_sum > 0 else s


def build_7x24_from_excel(df_load: pd.DataFrame, total: bool = False):
    """
    Convierte una tabla de carga en:
      - dow24: ndarray (7,24) con formas horarias por día (cada fila suma 1).
      - day_w: ndarray (7,) con pesos de energía semanal (suman 1).

    Columnas requeridas:
      - total=True: DOW, HOUR, TOTAL
      - total=False: DOW, HOUR, RES, IND, TOTAL
    """
    if total:
        req = {"DOW", "HOUR", "TOTAL"}
    else:
        req = {"DOW", "HOUR", "RES", "IND", "TOTAL"}
    if not req.issubset(set(df_load.columns)):
        expected = "DOW, HOUR, TOTAL" if total else "DOW, HOUR, RES, IND, TOTAL"
        raise ValueError(f"Profiles.LoadProfile debe tener: {expected}")

    df = df_load.copy()
    if not total:
        df["TOTAL"] = df["RES"].fillna(0) + df["IND"].fillna(0)

    dow24 = np.zeros((7, 24), dtype=float)
    day_energy = np.zeros(7, dtype=float)

    for d in range(1, 8):
        sel = df[df["DOW"] == d].sort_values("HOUR")
        v = sel["TOTAL"].to_numpy(dtype=float)
        if len(v) != 24:
            raise ValueError(f"DOW={d} no tiene 24 horas.")
        e = v.sum()
        if e <= 0:
            raise ValueError(f"Energía diaria nula para DOW={d}.")
        dow24[d - 1, :] = v / e
        day_energy[d - 1] = e

    day_w = day_energy / day_energy.sum()
    return dow24, day_w


DEFAULT_CONFIG = {
    "E_month_kWh": 2000.0,
    "use_excel_profile": True,
    "use_excel_seasonality": False,
    "alpha_mix": 0.5,
    "HSP": 5.5,
    "PR": 0.80,
    "panel_technology_mode": "standard",
    "deg_rate": 0.005,
    "Tmin_C": 10.0,
    "P_mod_W": 600.0,
    "Voc25": 50.0,
    "Vmp25": 41.0,
    "Isc": 14.0,
    "a_Voc_pct": -0.29,
    "ILR_min": 1.10,
    "ILR_max": 1.40,
    "buy_tariff_COP_kWh": 900.0,
    "sell_tariff_COP_kWh": 300.0,
    "g_tar_buy": 0.00,
    "g_tar_sell": 0.00,
    "discount_rate": 0.10,
    "years": 15,
    "pricing_mode": "variable",
    "price_per_kWp_COP": 3_500_000.0,
    "price_total_COP": 58_000_000.0,
    "include_var_others": False,
    "price_others_total": 0,
    "include_hw_in_price": False,
    "export_allowed": True,
    "include_battery": False,
    "battery_name": "",
    "optimize_battery": False,
    "bat_DoD": 0.90,
    "bat_eta_rt": 0.90,
    "kWp_seed_mode": "auto",
    "kWp_seed_manual_kWp": 15.0,
    "modules_span_each_side": 40,
    "kWp_min": 3.0,
    "kWp_max": 40.0,
    "limit_peak_ratio_enable": True,
    "limit_peak_ratio": 1.8,
    "limit_peak_year": 0,
    "limit_peak_month_mode": "max",
    "limit_peak_month_fixed": 1,
    "limit_peak_basis": "weighted_mean",
    "mc_PR_std": 0.07,
    "mc_buy_std": 0.06,
    "mc_sell_std": 0.10,
    "mc_demand_std": 0.04,
    "mc_use_manual_kWp": False,
    "mc_manual_kWp": 15.0,
    "mc_battery_name": "",
    "mc_n_simulations": 1000,
    "bat_coupling": "ac",
    "island_mode": False,
    "soc_week_steady": True,
    "soc_iter_max": 10,
    "soc_iter_tol_kWh": 1e-3,
}


def ensure_template(path: str):
    """Crea un Excel de entrada coherente con el contrato actual del loader."""
    from services.io_excel import ensure_template as ensure_template_service

    ensure_template_service(path)


def load_config_from_excel(path: str):
    """Carga configuración, perfiles y catálogos desde un Excel de entrada."""
    from services.io_excel import load_config_from_excel as load_config_service

    bundle = load_config_service(path)
    return (
        bundle.config,
        bundle.solar_profile,
        bundle.hsp_month,
        bundle.inverter_catalog,
        bundle.battery_catalog,
        bundle.demand_profile_7x24,
        bundle.day_weights,
        bundle.demand_month_factor,
        bundle.cop_kwp_table,
        bundle.cop_kwp_table_others,
    )


def _make_battery_obj(battery_dict):
    """Crea instancia Battery y SoC inicial (kWh) a partir de un dict de entrada."""
    if not battery_dict:
        return None, 0.0
    usable = (
        float(battery_dict["nominal_kWh"]) * float(battery_dict["DoD"])
        if ("nominal_kWh" in battery_dict and "DoD" in battery_dict)
        else float(battery_dict.get("usable_kWh", 0.0))
    )
    if usable <= 1e-9:
        return None, 0.0

    eta_ch = float(battery_dict.get("eta_ch", math.sqrt(battery_dict.get("eta_rt", 0.9))))
    eta_dis = float(battery_dict.get("eta_dis", math.sqrt(battery_dict.get("eta_rt", 0.9))))
    Pmax_ch = float(battery_dict.get("max_ch_kW", battery_dict.get("max_kW", 0.0)))
    Pmax_dis = float(battery_dict.get("max_dis_kW", battery_dict.get("max_kW", 0.0)))
    coupling = str(battery_dict.get("coupling", "ac")).strip().lower()
    SoC0 = battery_dict.get("SoC0", usable * 0.5)
    soc_kwh = SoC0 * usable if (0.0 <= float(SoC0) <= 1.0) else float(SoC0)
    soc_frac = soc_kwh / usable if usable > 0 else 0.5

    batt = Battery(
        name=str(battery_dict.get("name", "BAT")),
        usable_kwh=usable,
        max_ch_kw=Pmax_ch,
        max_dis_kw=Pmax_dis,
        eta_ch=eta_ch,
        eta_dis=eta_dis,
        coupling=coupling,
        soc_init=soc_frac,
        price_cop=float(battery_dict.get("price_COP", 0.0)),
    )
    return batt, soc_kwh


def _build_profiles(kWp, PR, HSP, s24, w24, E_day):
    """Genera perfiles de PV DC y demanda para un día dado."""
    load = np.array(w24, dtype=float) * float(E_day)
    E_pv_day = float(kWp) * float(PR) * float(HSP)
    pv_dc = np.array(s24, dtype=float) * E_pv_day
    return pv_dc, load


def _dispatch_one_day(kWp, PR, HSP, s24, w24, E_day, P_AC, battery, config_kwargs):
    """Wrapper sobre `dispatch_day` para devolver totales diarios resumidos."""
    pv_profile, load_profile = _build_profiles(kWp, PR, HSP, s24, w24, E_day)
    batt_obj, soc0 = _make_battery_obj(battery)
    cfg = DispatchConfig(inverter_ac_kw=float(P_AC), **config_kwargs)
    res = dispatch_day(pv_profile, load_profile, batt_obj, cfg, soc0)
    totals = res.day_totals()
    totals.update(
        dict(
            PV_gen_ac_day=float((res.pv_to_load + res.export_energy).sum()),
            Load_day=float(load_profile.sum()),
            SoC_end=float(res.soc_end),
            Unserved_day=float(res.unserved.sum()),
        )
    )
    return totals


def day_pv_load_flow_export(kWp, PR, HSP, s24, w24, E_day, P_AC, battery=None):
    """Despacho diario modo grid (import/export habilitado)."""
    return _dispatch_one_day(
        kWp,
        PR,
        HSP,
        s24,
        w24,
        E_day,
        P_AC,
        battery,
        config_kwargs=dict(allow_import=True, allow_export=True, export_limit_kw=None, mode="grid"),
    )


def day_pv_load_flow_zero_export(kWp, PR, HSP, s24, w24, E_day, P_AC, battery=None):
    """Despacho diario modo cero-inyección (permite import, limita export a 0)."""
    res = _dispatch_one_day(
        kWp,
        PR,
        HSP,
        s24,
        w24,
        E_day,
        P_AC,
        battery,
        config_kwargs=dict(allow_import=True, allow_export=True, export_limit_kw=0.0, mode="zero_export"),
    )
    res["Export_day"] = 0.0
    return res


def day_pv_load_flow_island(kWp, PR, HSP, s24, w24, E_day, P_AC, battery=None):
    """Despacho diario modo isla (sin import ni export; déficit se marca como no servido)."""
    res = _dispatch_one_day(
        kWp,
        PR,
        HSP,
        s24,
        w24,
        E_day,
        P_AC,
        battery,
        config_kwargs=dict(allow_import=False, allow_export=False, export_limit_kw=0.0, mode="island"),
    )
    res["Import_day"] = 0.0
    return res


def simulate_monthly_series_dow(
    cfg,
    kWp,
    inv_sel,
    battery_sel,
    export_allowed,
    years,
    dow24,
    day_w,
    s24,
    hsp_month,
    PR_month_std=0.0,
    buy_month_offset_std=0.0,
    sell_month_offset_std=0.0,
    demand_month_offset_std=0.0,
    rng=None,
    demand_month_factor=1,
    price_per_kwp_cop=None,
):
    """Simula series mensuales 7x24 reutilizando `Simulator` y retorna (df, summary)."""
    cfg_local = dict(cfg)
    cfg_local["mc_PR_std"] = PR_month_std
    cfg_local["mc_buy_std"] = buy_month_offset_std
    cfg_local["mc_sell_std"] = sell_month_offset_std
    cfg_local["mc_demand_std"] = demand_month_offset_std
    cfg_local["years"] = years

    system = PVSystem(
        kwp=kWp,
        pr=cfg_local["PR"],
        hsp_month=hsp_month,
        solar_shape=s24,
        inverter_ac_kw=inv_sel["inverter"]["AC_kW"],
        deg_rate=cfg_local.get("deg_rate", 0.0),
    )

    battery_obj = None
    if battery_sel is not None and float(battery_sel.get("nom_kWh", 0) or 0) > 0:
        usable = float(battery_sel["nom_kWh"]) * float(cfg_local["bat_DoD"])
        battery_obj = Battery(
            name=str(battery_sel.get("name", "BAT")),
            usable_kwh=usable,
            max_ch_kw=float(battery_sel.get("max_ch_kW", battery_sel.get("max_kW", 0.0))),
            max_dis_kw=float(battery_sel.get("max_dis_kW", battery_sel.get("max_kW", 0.0))),
            eta_ch=math.sqrt(float(cfg_local["bat_eta_rt"])),
            eta_dis=math.sqrt(float(cfg_local["bat_eta_rt"])),
            coupling=str(cfg_local.get("bat_coupling", "ac")).strip().lower(),
            soc_init=0.5,
            price_cop=float(battery_sel.get("price_COP", 0.0)),
        )

    island_mode = bool(cfg_local.get("island_mode", False))
    export_allowed_eff = False if island_mode else bool(export_allowed)
    dispatch_cfg = DispatchConfig(
        inverter_ac_kw=inv_sel["inverter"]["AC_kW"],
        allow_import=not island_mode,
        allow_export=export_allowed_eff,
        export_limit_kw=(0.0 if not export_allowed else None),
        mode=("island" if island_mode else ("zero_export" if not export_allowed else "grid")),
    )

    sim = Simulator(
        dow24=dow24,
        day_w=day_w,
        solar_shape=s24,
        hsp_month=hsp_month,
        demand_month_factor=demand_month_factor,
    )
    sim_res = sim.run(
        cfg=cfg_local,
        system=system,
        dispatch_cfg=dispatch_cfg,
        inv_sel=inv_sel,
        battery_sel=battery_sel,
        battery=battery_obj,
        years=years,
        price_per_kwp_cop=price_per_kwp_cop,
        rng=rng,
        stochastic=any(
            float(value or 0) > 0
            for value in (
                PR_month_std,
                buy_month_offset_std,
                sell_month_offset_std,
                demand_month_offset_std,
            )
        ),
    )
    return sim_res.monthly, sim_res.summary


def optimize_scan(
    cfg,
    inv_catalog,
    bat_catalog,
    dow24,
    day_w,
    s24,
    hsp_month: np.ndarray,
    export_allowed,
    demand_month_factor,
    cop_kwp_table,
    cop_kwp_table_others,
):
    """Ejecuta el escaneo de kWp/baterías devolviendo (best, scan_df, seed, detail)."""
    optimizer = Optimizer(
        cfg=cfg,
        inv_catalog=inv_catalog,
        bat_catalog=bat_catalog,
        dow24=dow24,
        day_w=day_w,
        s24=s24,
        hsp_month=hsp_month,
        demand_month_factor=demand_month_factor,
        export_allowed=export_allowed,
        cop_kwp_table=cop_kwp_table,
        cop_kwp_table_others=cop_kwp_table_others,
    )
    return optimizer.run()


def simulate_monte_carlo(
    cfg: dict,
    inv_df: pd.DataFrame,
    inv_sel,
    kWp_opt,
    best,
    bat_df: pd.DataFrame,
    cop_kwp_table,
    cop_kwp_table_others,
    export_allowed,
    dow24,
    s24,
    day_w,
    hsp_month,
    demand_month_factor,
):
    """Corre simulaciones Monte Carlo de payback usando el motor principal."""
    if bool(cfg.get("mc_use_manual_kWp", False)):
        n_mod = max(1, int(round(cfg["mc_manual_kWp"] * 1000.0 / cfg["P_mod_W"])))
        kWp_mc = n_mod * cfg["P_mod_W"] / 1000.0
        inv_sel_mc = inv_sel
        from .hardware import select_inverter_and_strings  # lazy import to avoid cycle

        inv_sel_manual = select_inverter_and_strings(
            kWp=kWp_mc,
            module=dict(P_mod_W=cfg["P_mod_W"], Voc25=cfg["Voc25"], Vmp25=cfg["Vmp25"], Isc=cfg["Isc"]),
            Tmin_C=cfg["Tmin_C"],
            a_Voc_pct=cfg["a_Voc_pct"],
            inv_catalog=inv_df,
            ILR_target=(cfg["ILR_min"], cfg["ILR_max"]),
        )
        if inv_sel_manual is not None:
            inv_sel_mc = inv_sel_manual
        batt_mc = best["battery"]
        if str(cfg.get("mc_battery_name", "")).strip() != "":
            row = bat_df[bat_df["name"].astype(str) == str(cfg["mc_battery_name"])]
            if not row.empty:
                batt_mc = row.iloc[0].to_dict()
    else:
        inv_sel_mc, kWp_mc, batt_mc = best["inv_sel"], best["kWp"], best["battery"]

    price_per_kwp_cop = cop_kwp_table.loc[
        (cop_kwp_table["MIN"] < kWp_mc) & (cop_kwp_table["MAX"] >= kWp_mc), "PRECIO_POR_KWP"
    ].values[0]
    price_per_kwp_cop_others = cop_kwp_table_others.loc[
        (cop_kwp_table_others["MIN"] < kWp_mc) & (cop_kwp_table_others["MAX"] >= kWp_mc), "PRECIO_POR_KWP"
    ].values[0]
    price_per_kwp_cop += price_per_kwp_cop_others

    rng = random.Random(123)
    pays = []
    for _ in range(int(cfg["mc_n_simulations"])):
        df_mc, summ_mc = simulate_monthly_series_dow(
            cfg=cfg,
            kWp=kWp_mc,
            inv_sel=inv_sel_mc,
            battery_sel=batt_mc,
            export_allowed=export_allowed,
            years=cfg["years"],
            dow24=dow24,
            day_w=day_w,
            s24=s24,
            hsp_month=hsp_month,
            PR_month_std=cfg["mc_PR_std"],
            buy_month_offset_std=cfg["mc_buy_std"],
            sell_month_offset_std=cfg["mc_sell_std"],
            demand_month_offset_std=cfg["mc_demand_std"],
            rng=rng,
            demand_month_factor=demand_month_factor,
            price_per_kwp_cop=price_per_kwp_cop,
        )
        pays.append(summ_mc["payback_years"] if summ_mc["payback_years"] is not None else np.nan)
    arr = np.array(pays, dtype=float)
    vals = arr[~np.isnan(arr)]
    return arr, vals


def plot_npv_scan(scan_df, seed_kwp, best_kwp, out_png, wp_panel):
    """Grafica NPV vs kWp y marca semilla/óptimo."""
    fig, ax = plt.subplots(figsize=(8.5, 4.8))
    g = scan_df.groupby("kWp", as_index=False)["NPV"].max()
    ax.plot(g["kWp"], g["NPV"], marker="o")
    ax.axvline(seed_kwp, linestyle="--", label=f"Semilla ~ {seed_kwp:.2f} kWp")
    ax.axvline(best_kwp, linestyle=":", label=f"Óptimo ~ {best_kwp:.2f} kWp")
    ax.set_xlabel("kWp instalado (múltiplos del módulo)")
    ax.set_ylabel("NPV (M-COP)")
    ax.grid(True, alpha=0.25)
    ax.legend()
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda val, _: f"{val/1e6:.0f}M"))
    secax = ax.secondary_xaxis("top", functions=(lambda x: x / (wp_panel / 1e3), lambda x: x * (wp_panel / 1e3)))
    secax.set_xlabel("Numero de Modulos Solares")
    fig.tight_layout()
    fig.savefig(out_png, dpi=160)
    plt.close(fig)


def prepare_autoconsumo_anual_series(df: pd.DataFrame, *, export_allowed: bool = True, lang: str = "es") -> dict:
    """Prepara series del gráfico anual de autoconsumo/importación/exportación (año 1)."""
    first_year = df.iloc[:12].copy()
    if "Año_mes" in first_year.columns:
        xlabels = [str(value) for value in first_year["Año_mes"].tolist()]
    else:
        xlabels = [str(index + 1) for index in range(len(first_year))]
    if {"PV_a_Carga_kWh", "Bateria_a_Carga_kWh", "Importacion_Red_kWh"}.issubset(first_year.columns):
        export = first_year.get("Exportacion_kWh", pd.Series(np.zeros(len(first_year))))
        pv_to_load = "PV → Carga" if lang == "es" else "PV to load"
        battery_to_load = "Batería → Carga" if lang == "es" else "Battery to load"
        grid_import = "Importación" if lang == "es" else "Grid import"
        export_label = "Exportación" if lang == "es" else "Export"
        series = [
            {"label": pv_to_load, "values": first_year["PV_a_Carga_kWh"].to_numpy(dtype=float), "color": "#57eb36"},
            {"label": battery_to_load, "values": first_year["Bateria_a_Carga_kWh"].to_numpy(dtype=float), "color": "#6fa8dc"},
            {"label": grid_import, "values": first_year["Importacion_Red_kWh"].to_numpy(dtype=float), "color": "#f26c4f"},
        ]
        if export_allowed and export is not None:
            series.append({"label": export_label, "values": export.to_numpy(dtype=float), "color": "#f7b32b"})
    else:
        autoconsumo = first_year.get("Autoconsumo_kWh", pd.Series(np.zeros(len(first_year))))
        imported = first_year.get("Demanda_Importada_no_Atendida_kWh", pd.Series(np.zeros(len(first_year))))
        self_consumption = "Autoconsumo" if lang == "es" else "Self-consumption"
        imported_demand = "Demanda importada" if lang == "es" else "Imported demand"
        export_label = "Exportación" if lang == "es" else "Export"
        series = [
            {"label": self_consumption, "values": autoconsumo.to_numpy(dtype=float), "color": "g"},
            {"label": imported_demand, "values": imported.to_numpy(dtype=float), "color": "red"},
        ]
        if export_allowed and "Exportacion_kWh" in first_year.columns:
            series.append({"label": export_label, "values": first_year["Exportacion_kWh"].to_numpy(dtype=float), "color": "orange"})
    title = (
        "Autoconsumo, Importación y Exportación (Año 1)"
        if lang == "es" and export_allowed
        else "Cobertura mensual de demanda (Año 1)"
        if lang == "es"
        else "Self-consumption, import, and export (Year 1)"
        if export_allowed
        else "Monthly demand coverage (Year 1)"
    )
    return {"xlabels": xlabels, "series": series, "title": title}


def plot_autoconsumo_anual(df: pd.DataFrame, out_dir: str, name_png: str, n_mods: int, export_allowed: bool = True, best: bool = False) -> None:
    """Grafica cobertura de demanda año 1 (PV directo, batería, import/export)."""
    prepared = prepare_autoconsumo_anual_series(df, export_allowed=export_allowed)
    plt.figure()
    x = np.arange(len(prepared["xlabels"]))
    bottom = np.zeros(len(prepared["xlabels"]), dtype=float)
    for series in prepared["series"]:
        values = np.asarray(series["values"], dtype=float)
        plt.bar(x, values, bottom=bottom, label=series["label"], color=series["color"])
        bottom = bottom + values

    props = dict(boxstyle="round", facecolor="whitesmoke", edgecolor="gray")
    ax = plt.gca()
    ax.text(0.01, 0.99, f"# módulos={int(n_mods)}", transform=ax.transAxes, fontsize=9, verticalalignment="top", horizontalalignment="left", bbox=props)

    plt.xticks(x, prepared["xlabels"], rotation=45)
    plt.xlabel("Mes")
    plt.ylabel("kWh")
    plt.title(prepared["title"])
    plt.legend()
    plt.tight_layout()
    if best:
        plt.savefig(os.path.join(out_dir, name_png), dpi=160)
    plt.savefig(os.path.join(os.path.join(out_dir, "detalle_bateria"), name_png), dpi=160)
    plt.close()


def _silverman_bandwidth(x):
    """Calcula ancho de banda de Silverman (robusto usando IQR)."""
    x = np.asarray(x, float)
    n = x.size
    if n < 2:
        return 1.0
    sigma = np.std(x, ddof=1)
    iqr = np.subtract(*np.percentile(x, [75, 25]))
    s = sigma if sigma > 0 else 1.0
    if iqr > 0:
        s = min(sigma, iqr / 1.34) if sigma > 0 else iqr / 1.34
    h = 0.9 * s * n ** (-1 / 5)
    return h if h > 1e-8 else max(1e-8, 0.2 * (np.max(x) - np.min(x) + 1e-12))


def _kde_gauss(x_grid, samples, h):
    """KDE gaussiana vectorizada (sin SciPy)."""
    samples = samples[None, :]
    xg = x_grid[:, None]
    z = (xg - samples) / h
    norm = np.exp(-0.5 * z ** 2) / np.sqrt(2 * np.pi)
    y = np.mean(norm, axis=1) / h
    return y


def plot_payback_kde(vals, out_dir, outfile="chart_payback_kde.png", title="Distribución de Payback (Monte Carlo)"):
    """Grafica KDE de payback con sombreado de cuantiles clave."""
    vals = np.asarray(vals, float)
    vals = vals[~np.isnan(vals)]
    if vals.size == 0:
        return None
    q80, q95 = np.percentile(vals, [80, 95])
    vmin, vmax = np.min(vals), np.max(vals)
    pad = 0.1 * (vmax - vmin + 1e-12)
    x_grid = np.linspace(vmin - pad, vmax + pad, 600)
    h = _silverman_bandwidth(vals)
    y = _kde_gauss(x_grid, vals, h) * 100
    plt.figure()
    plt.plot(x_grid, y, linewidth=2)
    plt.fill_between(x_grid, 0, y, where=(x_grid <= q80), alpha=0.35, color="green")
    plt.fill_between(x_grid, 0, y, where=(x_grid > q80) & (x_grid <= q95), alpha=0.35, color="yellow")
    plt.fill_between(x_grid, 0, y, where=(x_grid > q95), alpha=0.35, color="orange")
    plt.axvline(q80, linestyle="--", linewidth=1)
    plt.axvline(q95, linestyle="--", linewidth=1)
    plt.xlabel("Payback (años)")
    plt.ylabel("Probabilidad (%)")
    plt.title(title)
    plt.gca().yaxis.set_major_formatter(mticker.PercentFormatter())
    mean_val = np.mean(vals)
    median_val = np.median(vals)
    plt.text(
        0.98,
        0.95,
        f"Media ≈ {mean_val:.2f} años\nMediana ≈ {median_val:.2f} años",
        ha="right",
        va="top",
        transform=plt.gca().transAxes,
        fontsize=9,
        bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="gray", alpha=0.7),
    )
    plt.tight_layout()
    plt.savefig(os.path.join(out_dir, outfile), dpi=160)
    plt.close()
    return outfile


def prepare_typical_day_series(
    kWp,
    inv_sel,
    cfg,
    w24,
    s24,
    hsp_month,
    demand_month_factor,
    month_for_plot=None,
    year_for_plot=0,
    battery=None,
    export_allowed=True,
):
    """Prepara las series del gráfico de día típico."""
    if demand_month_factor is None:
        demand_month_factor = np.ones(12)
    if month_for_plot is None:
        month_for_plot = int(np.argmax(demand_month_factor))

    Dm = 30
    E_day = demand_month_factor[month_for_plot] * cfg["E_month_kWh"] / Dm
    P_AC = inv_sel["inverter"]["AC_kW"]
    pr_deg = (1.0 - cfg.get("deg_rate", 0.0)) ** year_for_plot
    PR_eff = cfg["PR"] * pr_deg

    demand_kw = np.asarray(w24, dtype=float) * E_day
    E_pv_day = kWp * PR_eff * hsp_month[month_for_plot]
    pv_dc = np.asarray(s24, dtype=float) * E_pv_day
    pv_ac = np.minimum(pv_dc, P_AC)
    battery_input = battery
    if battery and "usable_kWh" not in battery and "nom_kWh" in battery:
        battery_input = dict(battery)
        battery_input["usable_kWh"] = float(battery.get("nom_kWh", 0.0) or 0.0) * float(cfg.get("bat_DoD", 0.0) or 0.0)
    battery_obj, soc0 = _make_battery_obj(battery_input)
    export_limit_kw = None if bool(export_allowed) else 0.0
    dispatch = dispatch_day(
        pv_dc,
        demand_kw,
        battery_obj,
        DispatchConfig(
            inverter_ac_kw=float(P_AC),
            allow_import=True,
            allow_export=True,
            export_limit_kw=export_limit_kw,
            mode="grid" if bool(export_allowed) else "zero_export",
        ),
        soc0,
    )
    hours = np.arange(24)
    frame = pd.DataFrame(
        {
            "Hora": hours,
            "Demanda": demand_kw,
            "Autogeneración": pv_ac,
            "FV_a_Carga": dispatch.pv_to_load,
            "FV_a_Batería": dispatch.pv_to_batt,
            "Batería_a_Carga": dispatch.batt_to_load,
            "Importación_Red": dispatch.import_energy,
            "Exportación": dispatch.export_energy,
            "Recorte": dispatch.curtail,
            "Sol": np.asarray(s24, dtype=float),
        }
    )
    title = "Día Típico"
    return {
        "hours": hours,
        "demand_kw": demand_kw,
        "pv_ac_kw": pv_ac,
        "pv_to_load_kw": np.asarray(dispatch.pv_to_load, dtype=float),
        "pv_to_battery_kw": np.asarray(dispatch.pv_to_batt, dtype=float),
        "battery_to_load_kw": np.asarray(dispatch.batt_to_load, dtype=float),
        "grid_import_kw": np.asarray(dispatch.import_energy, dtype=float),
        "export_kw": np.asarray(dispatch.export_energy, dtype=float),
        "curtail_kw": np.asarray(dispatch.curtail, dtype=float),
        "has_battery": battery_obj is not None and float(battery_obj.usable_kwh) > 0.0,
        "solar_factor_pct": np.asarray(s24, dtype=float) * 100.0,
        "month_for_plot": int(month_for_plot),
        "title": title,
        "export_frame": frame,
    }


def plot_dia_tipico(
    kWp,
    inv_sel,
    cfg,
    w24,
    s24,
    export_allowed,
    out_path,
    out_dir,
    hsp_month,
    month_for_plot=None,
    year_for_plot=0,
    demand_month_factor=None,
    best=False,
    battery=None,
    name_png=None,
):
    """Dibuja el 'Día Típico' (consumo vs PV limitado por inversor)."""
    prepared = prepare_typical_day_series(
        kWp,
        inv_sel,
        cfg,
        w24,
        s24,
        hsp_month,
        demand_month_factor,
        month_for_plot=month_for_plot,
        year_for_plot=year_for_plot,
        battery=battery,
        export_allowed=export_allowed,
    )
    horas = prepared["hours"]
    fig, ax1 = plt.subplots(figsize=(9, 5))
    prepared["export_frame"].to_csv(os.path.join(out_dir, f"dia_tipico_{kWp}kWp.csv"), index=False)

    ax1.bar(horas - 0.2, prepared["demand_kw"], width=0.4, label="Consumo", color="red")
    ax1.bar(horas + 0.2, prepared["pv_ac_kw"], width=0.4, label="FV", color="#57eb36")
    if prepared.get("has_battery"):
        ax1.plot(horas, prepared["battery_to_load_kw"], linewidth=2.5, marker="o", label="Batería a carga", color="#2563eb")
        ax1.plot(horas, prepared["pv_to_battery_kw"], linewidth=2.0, marker="o", linestyle="--", label="FV a batería", color="#f59e0b")
        if np.any(np.abs(prepared["grid_import_kw"]) > 1e-6):
            ax1.plot(horas, prepared["grid_import_kw"], linewidth=1.8, linestyle=":", label="Importación red", color="#6b7280")

    props = dict(boxstyle="round", facecolor="whitesmoke", edgecolor="gray")
    ax1.text(0.01, 0.99, f"# módulos={int(1e3 * kWp / cfg['P_mod_W'])}", transform=ax1.transAxes, fontsize=9, verticalalignment="top", horizontalalignment="left", bbox=props)

    ax1.set_xlabel("Hora")
    ax1.set_ylabel("Potencia [kW]")
    ax1.set_xticks(horas)
    ax1.set_xlim(-0.5, 23.5)
    ax1.grid(axis="y", alpha=0.25)

    ax2 = ax1.twinx()
    ax2.plot(horas, prepared["solar_factor_pct"], linewidth=2.5, label="Factor Solar", color="#f2bb4b")
    ax2.set_ylabel("Factor Solar [%]")

    h1, l1 = ax1.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    ax1.legend(h1 + h2, l1 + l2, loc="upper right", ncol=3, frameon=True, framealpha=0.9)

    titulo = prepared["title"]
    if export_allowed is False:
        titulo += " (Cero inyección)"
    ax1.set_title(titulo)

    fig.tight_layout()
    if best:
        fig.savefig(out_path, dpi=160)
    if battery is not None:
        detail_dir = os.path.join(out_dir, "detalle_bateria")
        detail_name = name_png or os.path.basename(out_path)
        fig.savefig(os.path.join(detail_dir, detail_name), dpi=160)
    plt.close(fig)


def prepare_cumulative_npv_series(df: pd.DataFrame) -> dict:
    """Prepara series del gráfico de flujo descontado acumulado."""
    x_values = np.asarray(df["Año_mes"].values)
    y_values = np.asarray(df["NPV_COP"].values, dtype=float)
    colors = ["green" if value >= 0 else "red" for value in y_values]
    sign_changes = np.where(np.sign(y_values[:-1]) != np.sign(y_values[1:]))[0] if len(y_values) > 1 else np.asarray([], dtype=int)
    crossing_x = x_values[sign_changes[0] + 1] if len(sign_changes) > 0 else None
    return {
        "x_values": x_values,
        "y_values": y_values,
        "colors": colors,
        "crossing_x": crossing_x,
    }


def plot_cumulated_npv(df: pd.DataFrame, kWp, out_dir, cfg):
    """Grafica NPV acumulado mensual y marca el cruce a positivo si aplica."""
    prepared = prepare_cumulative_npv_series(df)
    fig, ax = plt.subplots()
    ax.bar(prepared["x_values"], prepared["y_values"], color=prepared["colors"])
    if prepared["crossing_x"] is not None:
        ax.axvline(prepared["crossing_x"], linestyle="--", color="blue", linewidth=1.5)

    plt.xticks(rotation=90)
    ax.xaxis.set_major_locator(mticker.MultipleLocator(10))
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda val, _: f"{val/1e6:.0f}M"))
    plt.ylabel("Valor Presente Neto acumulado (COP)")
    ax.axhline(0, linestyle="--", color="black")
    plt.title("Flujo descontado acumulado")
    props = dict(boxstyle="round", facecolor="whitesmoke", edgecolor="gray")
    ax.text(0.01, 0.99, f"# módulos={int(1e3 * kWp / cfg['P_mod_W'])}", transform=ax.transAxes, fontsize=9, verticalalignment="top", horizontalalignment="left", bbox=props)
    plt.tight_layout()
    plt.savefig(os.path.join(out_dir, f"grafica_flujo_acumulado_{kWp}kWp.png"), dpi=160)
    plt.close()


def prepare_battery_monthly_series(df_month: pd.DataFrame, *, lang: str = "es") -> dict:
    """Prepara series mensuales de cobertura de demanda y destino FV."""
    if "Año_mes" in df_month.columns:
        xlabels = df_month["Año_mes"].astype(str).tolist()
    else:
        xlabels = [str(i + 1) for i in range(len(df_month))]
    pv_load = df_month.get("PV_a_Carga_kWh", pd.Series(np.zeros(len(df_month)))).to_numpy(dtype=float)
    bat_load = df_month.get("Bateria_a_Carga_kWh", pd.Series(np.zeros(len(df_month)))).to_numpy(dtype=float)
    imp = df_month.get("Importacion_Red_kWh", pd.Series(np.zeros(len(df_month)))).to_numpy(dtype=float)
    pv_batt = df_month.get("PV_a_Bateria_kWh", pd.Series(np.zeros(len(df_month)))).to_numpy(dtype=float)
    exp = df_month.get("Exportacion_kWh", pd.Series(np.zeros(len(df_month)))).to_numpy(dtype=float)
    if "Curtailment_kWh" in df_month.columns:
        curtail = df_month["Curtailment_kWh"].to_numpy(dtype=float)
    else:
        curtail = np.zeros_like(exp)
    pv_to_load = "PV → Carga" if lang == "es" else "PV to load"
    battery_to_load = "Batería → Carga" if lang == "es" else "Battery to load"
    grid_import = "Importación Red" if lang == "es" else "Grid import"
    pv_to_battery = "PV → Batería" if lang == "es" else "PV to battery"
    export_label = "Exportación" if lang == "es" else "Export"
    curtailment = "Curtailment" if lang == "en" else "Recorte"
    return {
        "xlabels": xlabels,
        "coverage_series": [
            {"label": pv_to_load, "values": pv_load},
            {"label": battery_to_load, "values": bat_load},
            {"label": grid_import, "values": imp},
        ],
        "destination_series": [
            {"label": pv_to_load, "values": pv_load},
            {"label": pv_to_battery, "values": pv_batt},
            {"label": export_label, "values": exp},
            {"label": curtailment, "values": curtail},
        ],
    }


def plot_battery_monthly(df_month, kWp, cfg):
    """Grafica cobertura de demanda y destino PV mensual para batería (año 1)."""
    prepared = prepare_battery_monthly_series(df_month)
    xlabels = prepared["xlabels"]

    fig1, ax1 = plt.subplots(figsize=(12, 5))
    bottom = np.zeros(len(xlabels), dtype=float)
    for series in prepared["coverage_series"]:
        ax1.bar(xlabels, series["values"], bottom=bottom, label=series["label"])
        bottom = bottom + np.asarray(series["values"], dtype=float)
    ax1.set_title("Cobertura de la Demanda (mensual)")
    ax1.set_ylabel("Energía [kWh]")
    ax1.legend()
    ax1.grid(axis="y", linestyle=":")
    props = dict(boxstyle="round", facecolor="whitesmoke", edgecolor="gray")
    ax1.text(0.01, 0.99, f"# módulos={int(1e3 * kWp / cfg['P_mod_W'])}", transform=ax1.transAxes, fontsize=9, verticalalignment="top", horizontalalignment="left", bbox=props)

    fig2, ax2 = plt.subplots(figsize=(12, 5))
    bottom = np.zeros(len(xlabels), dtype=float)
    for series in prepared["destination_series"]:
        values = np.asarray(series["values"], dtype=float)
        if series["label"] == "Curtailment" and not np.any(values > 1e-6):
            continue
        ax2.bar(xlabels, values, bottom=bottom, label=series["label"])
        bottom = bottom + values

    ax2.set_title("Destino de la Generación FV (mensual)")
    ax2.set_ylabel("Energía [kWh]")
    ax2.legend()
    ax2.grid(axis="y", linestyle=":")
    ax2.text(0.01, 0.99, f"# módulos={int(1e3 * kWp / cfg['P_mod_W'])}", transform=ax2.transAxes, fontsize=9, verticalalignment="top", horizontalalignment="left", bbox=props)

    plt.tight_layout()
    return fig1, fig2
