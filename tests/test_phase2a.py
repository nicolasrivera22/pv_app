from __future__ import annotations

from dataclasses import replace
from io import BytesIO

import pandas.testing as pdt
from openpyxl import load_workbook
import pytest

from services.candidate_financials import (
    build_candidate_financial_snapshot,
    get_candidate_financial_snapshot_cache,
    resolve_financial_best_candidate_key,
)
from services import (
    ScenarioSessionState,
    add_scenario,
    apply_prepared_economics_runtime_price_bridge,
    build_comparison_figures,
    build_comparison_table,
    create_scenario_record,
    delete_scenario,
    duplicate_scenario,
    export_comparison_workbook,
    export_scenario_workbook,
    load_example_config,
    normalize_inverter_catalog_rows,
    prepare_economics_runtime_price_bridge,
    rename_scenario,
    run_scenario_scan,
    select_candidate_and_sync_runtime_price,
    set_comparison_scenarios,
    update_scenario_bundle,
    update_selected_candidate,
)
from services.economics_engine import EconomicsPreviewResult
from services.scenario_session import hydrate_scenario_scan, resolve_runtime_price_bridge_state


def _fast_bundle():
    bundle = load_example_config()
    config = {
        **bundle.config,
        "years": 5,
        "modules_span_each_side": 4,
        "kWp_min": 12.0,
        "kWp_max": 18.0,
    }
    return replace(bundle, config=config)


def _run_named_scenario(name: str, bundle=None):
    bundle = bundle or _fast_bundle()
    state = add_scenario(ScenarioSessionState.empty(), create_scenario_record(name, bundle))
    return run_scenario_scan(state, state.active_scenario_id)


def _bridge_selected_candidate(state: ScenarioSessionState) -> ScenarioSessionState:
    active = state.get_scenario()
    assert active is not None
    prepared = prepare_economics_runtime_price_bridge(active)
    assert prepared.applied is True
    return apply_prepared_economics_runtime_price_bridge(
        state,
        active.scenario_id,
        prepared,
        mark_project_dirty=False,
    )


def test_scenario_duplicate_rename_delete_workflow() -> None:
    state = add_scenario(ScenarioSessionState.empty(), create_scenario_record("Base", _fast_bundle()))
    active_id = state.active_scenario_id
    state = run_scenario_scan(state, active_id)

    duplicated = duplicate_scenario(state, active_id, new_name="Base copy")
    assert len(duplicated.scenarios) == 2
    assert duplicated.active_scenario_id != active_id
    assert duplicated.get_scenario().name == "Base copy"
    assert duplicated.get_scenario().scan_result is not None

    renamed = rename_scenario(duplicated, duplicated.active_scenario_id, "Variant A")
    assert renamed.get_scenario().name == "Variant A"

    reduced = delete_scenario(renamed, active_id)
    assert len(reduced.scenarios) == 1
    assert reduced.get_scenario().name == "Variant A"


def test_update_bundle_marks_scenario_dirty_and_clears_scan() -> None:
    state = _run_named_scenario("Editable")
    active = state.get_scenario()
    updated_bundle = replace(active.config_bundle, config={**active.config_bundle.config, "PR": 0.85})

    state = update_scenario_bundle(state, active.scenario_id, updated_bundle)
    active = state.get_scenario()

    assert active.dirty is True
    assert active.scan_result is None
    assert active.selected_candidate_key is None


def test_update_bundle_preserves_scan_for_economics_only_changes() -> None:
    state = _run_named_scenario("Economics only")
    active = state.get_scenario()
    assert active is not None
    assert active.scan_result is not None

    changed_costs = active.config_bundle.economics_cost_items_table.copy()
    changed_prices = active.config_bundle.economics_price_items_table.copy()
    changed_costs.at[0, "amount_COP"] = 123_456.0
    changed_prices.at[0, "value"] = 0.15
    updated_bundle = replace(
        active.config_bundle,
        economics_cost_items_table=changed_costs,
        economics_price_items_table=changed_prices,
    )

    state = update_scenario_bundle(state, active.scenario_id, updated_bundle)
    updated = state.get_scenario()

    assert updated is not None
    assert updated.dirty is False
    assert updated.scan_result is active.scan_result
    assert updated.scan_fingerprint == active.scan_fingerprint
    assert updated.selected_candidate_key == active.selected_candidate_key
    assert state.project_dirty is True


def test_run_scenario_scan_keeps_runtime_pricing_unchanged_until_manual_bridge() -> None:
    bundle = _fast_bundle()
    state = add_scenario(ScenarioSessionState.empty(), create_scenario_record("Auto bridge", bundle))

    state = run_scenario_scan(state, state.active_scenario_id)
    active = state.get_scenario()

    assert active is not None
    assert active.scan_result is not None
    assert active.runtime_price_bridge is None
    prepared = prepare_economics_runtime_price_bridge(active)
    assert prepared.applied is True
    assert active.selected_candidate_key == prepared.candidate_key
    assert active.config_bundle.config["pricing_mode"] == bundle.config["pricing_mode"]
    assert float(active.config_bundle.config["price_total_COP"]) == pytest.approx(float(bundle.config["price_total_COP"]))
    assert active.config_bundle.config["include_hw_in_price"] == bundle.config["include_hw_in_price"]
    assert float(active.config_bundle.config["price_others_total"]) == pytest.approx(float(bundle.config["price_others_total"]))


def test_update_selected_candidate_remains_selection_only() -> None:
    state = _run_named_scenario("Selection primitive")
    active = state.get_scenario()
    assert active is not None
    assert active.scan_result is not None
    assert active.runtime_price_bridge is None
    non_best_key = next(
        key for key in active.scan_result.candidate_details if key != active.selected_candidate_key
    )
    original_price_total = float(active.config_bundle.config["price_total_COP"])
    original_bridge = active.runtime_price_bridge

    state = update_selected_candidate(state, active.scenario_id, non_best_key)
    updated = state.get_scenario()

    assert updated is not None
    assert updated.selected_candidate_key == non_best_key
    assert float(updated.config_bundle.config["price_total_COP"]) == pytest.approx(original_price_total)
    assert updated.runtime_price_bridge == original_bridge


def test_select_candidate_and_sync_runtime_price_updates_runtime_fields() -> None:
    state = _run_named_scenario("Selection sync")
    active = state.get_scenario()
    assert active is not None
    assert active.scan_result is not None
    non_best_key = None
    expected = None
    for candidate_key in active.scan_result.candidate_details:
        if candidate_key == active.selected_candidate_key:
            continue
        prepared = prepare_economics_runtime_price_bridge(replace(active, selected_candidate_key=candidate_key))
        if prepared.applied:
            non_best_key = candidate_key
            expected = prepared
            break
    assert non_best_key is not None
    assert expected is not None
    assert expected.applied is True

    state, prepared = select_candidate_and_sync_runtime_price(state, active.scenario_id, non_best_key)
    updated = state.get_scenario()

    assert prepared.applied is True
    assert updated is not None
    assert updated.selected_candidate_key == non_best_key
    assert updated.config_bundle.config["pricing_mode"] == "total"
    assert float(updated.config_bundle.config["price_total_COP"]) == pytest.approx(float(expected.final_price_COP))
    assert updated.config_bundle.config["include_hw_in_price"] is False
    assert float(updated.config_bundle.config["price_others_total"]) == pytest.approx(0.0)
    assert updated.runtime_price_bridge is not None
    assert updated.runtime_price_bridge.candidate_key == non_best_key
    assert updated.runtime_price_bridge.applied_economics_signature is not None


def test_hydrate_scenario_scan_does_not_auto_bridge() -> None:
    state = _run_named_scenario("Hydrate only")
    active = state.get_scenario()
    assert active is not None
    stripped = replace(active, scan_result=None, runtime_price_bridge=None)
    state = replace(
        state,
        scenarios=tuple(stripped if item.scenario_id == stripped.scenario_id else item for item in state.scenarios),
    )

    hydrated = hydrate_scenario_scan(state, active.scenario_id)
    hydrated_active = hydrated.get_scenario()

    assert hydrated_active is not None
    assert hydrated_active.scan_result is not None
    assert hydrated_active.runtime_price_bridge is None


def test_selected_candidate_resets_after_scenario_change_and_rerun() -> None:
    state = _run_named_scenario("Selection test")
    active = state.get_scenario()
    assert active.scan_result is not None
    non_best_key = next(
        key
        for key, detail in active.scan_result.candidate_details.items()
        if key != active.scan_result.best_candidate_key and float(detail["kWp"]) >= 16.0
    )

    state = update_selected_candidate(state, active.scenario_id, non_best_key)
    selected = state.get_scenario().selected_candidate_key
    assert selected == non_best_key

    narrowed_bundle = replace(
        active.config_bundle,
        config={**active.config_bundle.config, "kWp_max": 13.8},
    )
    state = update_scenario_bundle(state, active.scenario_id, narrowed_bundle)
    state = run_scenario_scan(state, active.scenario_id)
    rerun = state.get_scenario()

    assert rerun.selected_candidate_key == resolve_financial_best_candidate_key(rerun)
    assert non_best_key not in rerun.scan_result.candidate_details


def test_blocked_selection_sync_keeps_same_candidate_bridge_active(monkeypatch) -> None:
    state = _run_named_scenario("Blocked same candidate")
    state = _bridge_selected_candidate(state)
    active = state.get_scenario()
    assert active is not None
    assert active.runtime_price_bridge is not None

    def _blocked_preview(_scenario, *, economics_cost_items, economics_price_items):
        _ = economics_cost_items, economics_price_items
        return EconomicsPreviewResult(
            state="rerun_required",
            message_key="workspace.admin.economics.preview.state.rerun_required",
        )

    monkeypatch.setattr("services.scenario_session.resolve_economics_preview", _blocked_preview)

    next_state, prepared = select_candidate_and_sync_runtime_price(
        state,
        active.scenario_id,
        active.selected_candidate_key,
    )
    next_active = next_state.get_scenario()

    assert prepared.applied is False
    assert next_active is not None
    assert next_active.runtime_price_bridge is not None
    assert next_active.runtime_price_bridge.stale is False
    assert resolve_runtime_price_bridge_state(next_active) == "active"


def test_blocked_selection_sync_stales_bridge_when_governing_design_changes(monkeypatch) -> None:
    state = _run_named_scenario("Blocked other candidate")
    state = _bridge_selected_candidate(state)
    active = state.get_scenario()
    assert active is not None
    assert active.scan_result is not None
    assert active.runtime_price_bridge is not None
    original_price_total = float(active.config_bundle.config["price_total_COP"])
    non_best_key = next(
        key for key in active.scan_result.candidate_details if key != active.scan_result.best_candidate_key
    )

    def _blocked_preview(_scenario, *, economics_cost_items, economics_price_items):
        _ = economics_cost_items, economics_price_items
        return EconomicsPreviewResult(
            state="rerun_required",
            message_key="workspace.admin.economics.preview.state.rerun_required",
        )

    monkeypatch.setattr("services.scenario_session.resolve_economics_preview", _blocked_preview)

    next_state, prepared = select_candidate_and_sync_runtime_price(state, active.scenario_id, non_best_key)
    next_active = next_state.get_scenario()

    assert prepared.applied is False
    assert next_active is not None
    assert next_active.selected_candidate_key == non_best_key
    assert float(next_active.config_bundle.config["price_total_COP"]) == pytest.approx(original_price_total)
    assert next_active.runtime_price_bridge is not None
    assert next_active.runtime_price_bridge.stale is True
    assert resolve_runtime_price_bridge_state(next_active) == "stale"


def test_catalog_row_validation_reports_duplicate_names_and_nonnumeric_values() -> None:
    rows = [
        {"name": "INV-A", "AC_kW": "10", "Vmppt_min": 200, "Vmppt_max": 800, "Vdc_max": 1000, "Imax_mppt": 18, "n_mppt": 2, "price_COP": 9_000_000},
        {"name": "INV-A", "AC_kW": "banana", "Vmppt_min": 200, "Vmppt_max": 800, "Vdc_max": 1000, "Imax_mppt": 18, "n_mppt": 2, "price_COP": 9_000_000},
    ]

    _, issues = normalize_inverter_catalog_rows(rows)

    messages = [issue.message for issue in issues]
    assert any("Duplicados" in message for message in messages)
    assert any("debe ser numérico" in message for message in messages)


def test_comparison_outputs_are_stable_for_two_scenarios() -> None:
    state = add_scenario(ScenarioSessionState.empty(), create_scenario_record("Base", _fast_bundle()))
    variant_bundle = replace(_fast_bundle(), config={**_fast_bundle().config, "buy_tariff_COP_kWh": 1050.0})
    state = add_scenario(state, create_scenario_record("Variant", variant_bundle))

    scenario_ids = [scenario.scenario_id for scenario in state.scenarios]
    for scenario_id in scenario_ids:
        state = run_scenario_scan(state, scenario_id)
    state = set_comparison_scenarios(state, scenario_ids)

    rows = [state.get_scenario(scenario_id) for scenario_id in state.comparison_scenario_ids]
    table_a = build_comparison_table(rows)
    table_b = build_comparison_table(rows)
    pdt.assert_frame_equal(table_a, table_b)

    figures = build_comparison_figures(rows)
    assert len(figures["npv_overlay"].data) == 2
    assert len(figures["kpi_bar"].data) == 3


def test_export_scenario_workbook_contains_expected_sheets() -> None:
    state = _run_named_scenario("Export")
    scenario = state.get_scenario()
    payload = export_scenario_workbook(scenario)

    workbook = load_workbook(BytesIO(payload))
    assert workbook.sheetnames == ["Summary", "Config", "Inverters", "Batteries", "Candidates", "Monthly_Selected"]


def test_export_comparison_workbook_contains_expected_sheets() -> None:
    state = add_scenario(ScenarioSessionState.empty(), create_scenario_record("Base", _fast_bundle()))
    variant_bundle = replace(_fast_bundle(), config={**_fast_bundle().config, "sell_tariff_COP_kWh": 350.0})
    state = add_scenario(state, create_scenario_record("Variant", variant_bundle))
    for scenario in state.scenarios:
        state = run_scenario_scan(state, scenario.scenario_id)
    state = set_comparison_scenarios(state, [scenario.scenario_id for scenario in state.scenarios])

    selected = [state.get_scenario(scenario_id) for scenario_id in state.comparison_scenario_ids]
    payload = export_comparison_workbook(state, selected)

    workbook = load_workbook(BytesIO(payload))
    assert "Comparison_Summary" in workbook.sheetnames
    assert "Comparison_KPIs" in workbook.sheetnames
    candidate_sheets = [name for name in workbook.sheetnames if name.startswith("Candidates_")]
    assert len(candidate_sheets) == 2


def test_candidate_financial_snapshot_cache_reuses_key_and_invalidates_on_economics_change() -> None:
    get_candidate_financial_snapshot_cache().clear()
    state = _run_named_scenario("Snapshot cache")
    active = state.get_scenario()
    assert active is not None
    candidate_key = active.selected_candidate_key
    assert candidate_key is not None

    first = build_candidate_financial_snapshot(active, candidate_key)
    second = build_candidate_financial_snapshot(active, candidate_key)
    assert first is second

    changed_prices = active.config_bundle.economics_price_items_table.copy()
    changed_prices.at[0, "value"] = float(changed_prices.at[0, "value"]) + 0.05
    changed_bundle = replace(active.config_bundle, economics_price_items_table=changed_prices)
    changed_state = update_scenario_bundle(state, active.scenario_id, changed_bundle)
    changed_active = changed_state.get_scenario()
    assert changed_active is not None

    third = build_candidate_financial_snapshot(changed_active, candidate_key)
    assert third is not first
    assert third.economics_signature != first.economics_signature
    assert third.visible_npv_COP != pytest.approx(first.visible_npv_COP)


def test_export_scenario_workbook_uses_snapshot_finance_contract() -> None:
    state = _run_named_scenario("Export snapshot")
    scenario = state.get_scenario()
    assert scenario is not None
    candidate_key = scenario.selected_candidate_key
    assert candidate_key is not None
    snapshot = build_candidate_financial_snapshot(scenario, candidate_key)

    payload = export_scenario_workbook(scenario)
    workbook = load_workbook(BytesIO(payload), data_only=True)
    summary_sheet = workbook["Summary"]
    summary_values = {
        summary_sheet.cell(row=row_index, column=1).value: summary_sheet.cell(row=row_index, column=2).value
        for row_index in range(2, summary_sheet.max_row + 1)
    }
    monthly_sheet = workbook["Monthly_Selected"]
    headers = [monthly_sheet.cell(row=1, column=column_index).value for column_index in range(1, monthly_sheet.max_column + 1)]
    npv_column = headers.index("NPV_COP") + 1
    last_npv = monthly_sheet.cell(row=monthly_sheet.max_row, column=npv_column).value

    assert summary_values["candidate_key"] == candidate_key
    assert float(summary_values["capex_client_COP"]) == pytest.approx(snapshot.capex_client_COP)
    assert float(summary_values["NPV_COP"]) == pytest.approx(snapshot.visible_npv_COP)
    assert float(last_npv) == pytest.approx(snapshot.visible_npv_COP)
