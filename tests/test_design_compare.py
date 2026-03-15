from __future__ import annotations

from dataclasses import replace
from io import BytesIO

from openpyxl import load_workbook

from services import (
    ScenarioSessionState,
    add_scenario,
    create_scenario_record,
    export_design_comparison_workbook,
    load_example_config,
    run_scenario_scan,
    set_design_comparison_candidates,
    update_selected_candidate,
)
from services.design_compare import (
    MAX_COMPARE_DESIGNS,
    append_design_selection,
    build_available_design_rows,
    build_design_compare_state,
    build_design_comparison_figures,
    build_design_comparison_rows,
    build_monthly_pv_destination_figure,
    build_monthly_pv_destination_frame,
    build_npv_projection_figure,
    build_npv_projection_frame,
    build_typical_day_figure,
    derive_panel_count,
    remove_design_selection,
    resolve_design_selection,
)


def _fast_bundle():
    bundle = load_example_config()
    config = {
        **bundle.config,
        "years": 5,
        "modules_span_each_side": 4,
        "kWp_min": 12.0,
        "kWp_max": 24.6,
    }
    return replace(bundle, config=config)


def _run_ready_state() -> ScenarioSessionState:
    state = add_scenario(ScenarioSessionState.empty(), create_scenario_record("Caso base", _fast_bundle()))
    return run_scenario_scan(state, state.active_scenario_id)


def test_design_selection_defaults_to_active_workbench_candidate() -> None:
    state = _run_ready_state()
    scenario = state.get_scenario()
    assert scenario is not None and scenario.scan_result is not None

    non_best = next(key for key in scenario.scan_result.candidate_details if key != scenario.scan_result.best_candidate_key)
    state = update_selected_candidate(state, scenario.scenario_id, non_best)
    resolved = resolve_design_selection(state, state.get_scenario())

    assert resolved == (non_best,)


def test_explicit_empty_selection_overrides_default() -> None:
    state = _run_ready_state()
    scenario = state.get_scenario()
    assert scenario is not None

    state = set_design_comparison_candidates(state, scenario.scenario_id, ())
    resolved = resolve_design_selection(state, state.get_scenario())

    assert resolved == ()


def test_selection_append_dedupes_and_caps_at_ten() -> None:
    state = _run_ready_state()
    scenario = state.get_scenario()
    assert scenario is not None and scenario.scan_result is not None

    keys = list(scenario.scan_result.candidate_details)[:12]
    appended = append_design_selection(scenario, (), keys)
    repeated = append_design_selection(scenario, appended, keys[:3])

    assert len(appended) == MAX_COMPARE_DESIGNS
    assert repeated == appended
    assert len(set(appended)) == len(appended)


def test_panel_count_prefers_n_mod_and_falls_back_to_integer_rounding() -> None:
    state = _run_ready_state()
    scenario = state.get_scenario()
    assert scenario is not None and scenario.scan_result is not None
    detail = scenario.scan_result.candidate_details[scenario.scan_result.best_candidate_key]

    assert derive_panel_count(detail, scenario) == int(detail["inv_sel"]["N_mod"])

    fallback_detail = {
        **detail,
        "inv_sel": {**detail["inv_sel"], "N_mod": None},
        "kWp": 12.0,
    }
    scenario = replace(scenario, config_bundle=replace(scenario.config_bundle, config={**scenario.config_bundle.config, "P_mod_W": 600.0}))
    assert derive_panel_count(fallback_detail, scenario) == 20


def test_design_rows_include_customer_facing_fields() -> None:
    state = _run_ready_state()
    scenario = state.get_scenario()
    assert scenario is not None and scenario.scan_result is not None
    keys = list(scenario.scan_result.candidate_details)[:2]

    rows = build_design_comparison_rows(scenario, keys, lang="es")
    available = build_available_design_rows(scenario, lang="es")

    assert list(rows["design_label"]) == ["#1", "#2"]
    assert "panel_count" in rows.columns
    assert rows["panel_count"].map(lambda value: int(value) == value).all()
    assert "inverter_name" in rows.columns
    assert "candidate_key" in available.columns
    assert "is_workbench_selected" in available.columns


def test_selected_design_rows_keep_first_selection_and_sort_remaining_by_kwp_then_battery() -> None:
    state = _run_ready_state()
    scenario = state.get_scenario()
    assert scenario is not None and scenario.scan_result is not None

    keys = list(scenario.scan_result.candidate_details)
    pinned = keys[-1]
    remainder = keys[:4]
    selection = (pinned, *remainder)

    rows = build_design_comparison_rows(scenario, selection, lang="es")

    expected_remainder = sorted(
        remainder,
        key=lambda candidate_key: (
            float(scenario.scan_result.candidate_details[candidate_key]["kWp"]),
            str(scenario.scan_result.candidate_details[candidate_key]["battery_name"]),
            candidate_key,
        ),
    )
    assert rows["candidate_key"].tolist() == [pinned, *expected_remainder]
    assert rows["design_label"].tolist() == [f"#{index}" for index in range(1, len(rows) + 1)]


def test_compare_state_distinguishes_blocked_and_ready_modes() -> None:
    empty_state = build_design_compare_state(None, (), lang="es")
    assert empty_state.code == "no_active"

    state = add_scenario(ScenarioSessionState.empty(), create_scenario_record("Nuevo", _fast_bundle()))
    scenario = state.get_scenario()
    assert scenario is not None
    no_scan = build_design_compare_state(scenario, (), lang="es")
    assert no_scan.code == "no_scan"

    ready_state = _run_ready_state()
    ready_scenario = ready_state.get_scenario()
    ready_selection = resolve_design_selection(ready_state, ready_scenario)
    ready = build_design_compare_state(ready_scenario, ready_selection, lang="es")
    assert ready.code == "ready"
    assert ready.can_select is True
    assert ready.can_export is False

    ready_state = set_design_comparison_candidates(
        ready_state,
        ready_scenario.scenario_id,
        list(ready_scenario.scan_result.candidate_details)[:2],
    )
    ready_scenario = ready_state.get_scenario()
    selected = resolve_design_selection(ready_state, ready_scenario)
    export_ready = build_design_compare_state(ready_scenario, selected, lang="es")
    assert export_ready.can_export is True


def test_monthly_destination_chart_switches_layout_for_large_selection() -> None:
    state = _run_ready_state()
    scenario = state.get_scenario()
    assert scenario is not None and scenario.scan_result is not None

    small_keys = list(scenario.scan_result.candidate_details)[:3]
    large_keys = list(scenario.scan_result.candidate_details)[:5]

    small_frame = build_monthly_pv_destination_frame(scenario, small_keys)
    large_frame = build_monthly_pv_destination_frame(scenario, large_keys)

    small_figure = build_monthly_pv_destination_figure(small_frame, lang="es", empty_message="vacío")
    large_figure = build_monthly_pv_destination_figure(large_frame, lang="es", empty_message="vacío")

    assert small_figure.layout.barmode == "stack"
    assert list(small_figure.layout.xaxis.ticktext) == ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    first_design_pv_trace = next(trace for trace in small_figure.data if trace.name.startswith("#1 · FV a carga"))
    second_design_pv_trace = next(trace for trace in small_figure.data if trace.name.startswith("#2 · FV a carga"))
    assert first_design_pv_trace.marker.color == "#16a34a"
    assert second_design_pv_trace.marker.color != first_design_pv_trace.marker.color
    assert len(large_figure.layout.annotations or []) >= len(large_keys)
    assert list(large_figure.layout.xaxis.ticktext) == ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


def test_typical_day_figure_uses_small_multiples_and_dynamic_height() -> None:
    state = _run_ready_state()
    scenario = state.get_scenario()
    assert scenario is not None and scenario.scan_result is not None
    keys = list(scenario.scan_result.candidate_details)[:5]

    figures = build_design_comparison_figures(scenario, keys, lang="es", empty_message="vacío")
    typical_day = figures["typical_day"]

    assert len(typical_day.layout.annotations or []) >= len(keys)
    assert int(typical_day.layout.height) >= 640


def test_remove_design_selection_and_export_workbook() -> None:
    state = _run_ready_state()
    scenario = state.get_scenario()
    assert scenario is not None and scenario.scan_result is not None
    keys = list(scenario.scan_result.candidate_details)[:3]

    removed = remove_design_selection(scenario, keys, keys[1])
    assert removed == (keys[0], keys[2])

    payload = export_design_comparison_workbook(scenario, keys[:2], lang="es")
    workbook = load_workbook(BytesIO(payload))

    assert workbook.sheetnames == [
        "Design_Comparison_Summary",
        "Design_Comparison_Metrics",
        "Annual_Demand_Coverage",
        "Monthly_PV_Destination",
        "Typical_Day",
        "NPV_Projection",
    ]


def test_npv_projection_frame_and_figure_use_calendar_and_project_year_axes() -> None:
    state = _run_ready_state()
    scenario = state.get_scenario()
    assert scenario is not None and scenario.scan_result is not None
    keys = list(scenario.scan_result.candidate_details)[:2]

    frame = build_npv_projection_frame(scenario, keys, lang="es", base_year=2026)
    figure = build_npv_projection_figure(frame, lang="es", empty_message="vacío", base_year=2026)

    assert {"month_index", "calendar_year", "project_year"}.issubset(frame.columns)
    assert frame.iloc[0]["calendar_year"] == 2027
    assert frame.iloc[0]["project_year"] == 1
    assert figure.layout.xaxis.title.text == "Año calendario"
    assert figure.layout.xaxis2.title.text == "Horizonte del proyecto"
    assert figure.layout.xaxis2.ticktext[0] == "Año 1"
    assert len(figure.data) == len(keys)
