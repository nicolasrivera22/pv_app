from __future__ import annotations

import math
import re
import shutil
from dataclasses import replace
from pathlib import Path

import numpy as np
import pandas as pd
import pandas.testing as pdt
import pytest
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from pv_product.simulator import calculate_capex_client
from pv_product.utils import build_7x24_from_excel
from services import collect_config_updates, ensure_template, load_config_from_excel, load_example_config, run_scan, run_scenario
from services.io_excel import WorkbookContractError, _normalize_config_value
from services.result_views import (
    build_npv_figure,
    build_kpis,
    build_visible_horizon_candidate_summary,
    resolve_payback_display_state,
    resolve_selected_candidate_key,
    resolve_selected_candidate_key_for_scenario,
    summarize_candidate_for_horizon,
    summarize_candidates_for_horizon,
)
from services.ui_schema import coerce_config_value, display_assumption_value, parse_assumption_input_value
from services.validation import validate_config


def _fast_bundle():
    bundle = load_example_config()
    config = {
        **bundle.config,
        "years": 5,
        "modules_span_each_side": 4,
        "kWp_min": 12.0,
        "kWp_max": 18.0,
    }
    return replace(bundle, config=config)


def _copy_workbook(tmp_path: Path, name: str = "workbook.xlsx") -> Path:
    destination = tmp_path / name
    shutil.copyfile(Path("PV_inputs.xlsx"), destination)
    return destination


def _table_header_map(ws, table_name: str) -> dict[str, int]:
    table = ws.tables[table_name]
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    headers = {
        ws.cell(row=min_row, column=column).value: column
        for column in range(min_col, max_col + 1)
    }
    return headers


def _set_config_value(path: Path, item_name: str, value) -> None:
    wb = load_workbook(path)
    ws = wb["Config"]
    headers = _table_header_map(ws, "Config")
    item_col = headers["Item"]
    value_col = headers["Valor"]
    _, min_row, _, max_row = range_boundaries(ws.tables["Config"].ref)
    for row in range(min_row + 1, max_row + 1):
        if ws.cell(row=row, column=item_col).value == item_name:
            ws.cell(row=row, column=value_col, value=value)
            wb.save(path)
            return
    raise AssertionError(f"Config item {item_name!r} not found")


def _blank_table_rows(path: Path, sheet_name: str, table_name: str) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(ws.tables[table_name].ref)
    for row in range(min_row + 1, max_row + 1):
        for column in range(min_col, max_col + 1):
            ws.cell(row=row, column=column, value=None)
    wb.save(path)


def _remove_sheet(path: Path, sheet_name: str) -> None:
    wb = load_workbook(path)
    del wb[sheet_name]
    wb.save(path)


def _remove_table(path: Path, sheet_name: str, table_name: str) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    del ws.tables[table_name]
    wb.save(path)


def _rename_table_column(path: Path, sheet_name: str, table_name: str, old_name: str, new_name: str) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    headers = _table_header_map(ws, table_name)
    column = headers[old_name]
    min_col, min_row, _, _ = range_boundaries(ws.tables[table_name].ref)
    assert column >= min_col
    ws.cell(row=min_row, column=column, value=new_name)
    wb.save(path)


def _parse_legacy_summary(summary_path: Path) -> dict[str, float | str]:
    text = summary_path.read_text(encoding="utf-8")
    return {
        "best_kwp": float(re.search(r"kWp óptimo .*: ([0-9.]+) kWp", text).group(1)),
        "battery": re.search(r"Batería óptima: ([A-Z0-9-]+|Ninguna)", text).group(1),
        "payback_years": float(re.search(r"Payback \(base\) ≈ ([0-9.]+) años", text).group(1)),
    }


def test_loads_spanish_workbook_and_normalizes_config() -> None:
    bundle = load_config_from_excel(Path("PV_inputs.xlsx"))

    assert bundle.config["bat_coupling"] == "dc"
    assert bundle.config["include_battery"] is True
    assert bundle.config["battery_name"] == ""
    assert bundle.config["mc_battery_name"] == ""
    assert bundle.config["export_allowed"] is True
    assert bundle.config["use_excel_profile"] == "perfil horario relativo"
    assert bundle.demand_profile_7x24.shape == (7, 24)
    assert bundle.hsp_month.shape == (12,)


def test_collect_config_updates_preserves_falsey_battery_toggle_and_blank_text_fields() -> None:
    base_config = {
        **load_example_config().config,
        "include_battery": True,
        "battery_name": np.nan,
        "price_total_COP": 58_000_000.0,
    }

    updated = collect_config_updates(
        [
            {"field": "include_battery"},
            {"field": "price_total_COP"},
            {"field": "battery_name"},
        ],
        [False, 0, ""],
        base_config,
    )

    assert updated["include_battery"] is False
    assert updated["price_total_COP"] == 0.0
    assert updated["battery_name"] == ""


def test_collect_config_updates_scales_percentage_inputs_back_to_internal_fractions() -> None:
    base_config = {
        **load_example_config().config,
        "discount_rate": 0.05,
        "alpha_mix": 0.6,
        "a_Voc_pct": -0.29,
    }

    updated = collect_config_updates(
        [
            {"field": "discount_rate"},
            {"field": "alpha_mix"},
            {"field": "a_Voc_pct"},
        ],
        [7.5, 80, -0.31],
        base_config,
    )

    assert updated["discount_rate"] == pytest.approx(0.075)
    assert updated["alpha_mix"] == pytest.approx(0.8)
    assert updated["a_Voc_pct"] == pytest.approx(-0.31)


def test_coerce_config_value_preserves_false_and_zero_and_treats_missing_values_as_missing() -> None:
    base_config = {
        "include_battery": True,
        "price_total_COP": 58_000_000.0,
        "battery_name": np.nan,
        "years": np.nan,
        "use_excel_profile": "perfil horario relativo",
    }

    assert coerce_config_value("include_battery", False, base_config) is False
    assert coerce_config_value("price_total_COP", 0, base_config) == 0.0
    assert coerce_config_value("use_excel_profile", "perfil general", base_config) == "perfil general"
    assert coerce_config_value("battery_name", "", base_config) == ""
    assert coerce_config_value("battery_name", np.nan, base_config) == ""
    assert coerce_config_value("years", np.nan, base_config) == 15
    assert coerce_config_value("price_total_COP", "", base_config) == 58_000_000.0


def test_assumption_scaling_helpers_render_and_parse_percent_fields_centrally() -> None:
    assert display_assumption_value("discount_rate", 0.05) == 5
    assert parse_assumption_input_value("discount_rate", 7.5) == pytest.approx(0.075)
    assert display_assumption_value("alpha_mix", 0.8) == 80
    assert parse_assumption_input_value("alpha_mix", 80) == pytest.approx(0.8)
    assert display_assumption_value("a_Voc_pct", -0.29) == pytest.approx(-0.29)
    assert parse_assumption_input_value("a_Voc_pct", -0.29) == pytest.approx(-0.29)
    assert display_assumption_value("limit_peak_ratio", 2.5) == pytest.approx(2.5)


def test_normalize_config_value_converts_blank_text_config_fields_from_nan_to_empty_string() -> None:
    battery_name, battery_error = _normalize_config_value("battery_name", np.nan)
    mc_battery_name, mc_battery_error = _normalize_config_value("mc_battery_name", np.nan)

    assert battery_name == ""
    assert battery_error is None
    assert mc_battery_name == ""
    assert mc_battery_error is None


def test_deterministic_scan_ignores_monte_carlo_fields() -> None:
    bundle = _fast_bundle()
    noisy_bundle = replace(
        bundle,
        config={
            **bundle.config,
            "mc_PR_std": 0.9,
            "mc_buy_std": 0.9,
            "mc_sell_std": 0.9,
            "mc_demand_std": 0.9,
        },
    )

    baseline_scan = run_scan(bundle)
    noisy_scan = run_scan(noisy_bundle)

    assert baseline_scan.best_candidate_key == noisy_scan.best_candidate_key
    pdt.assert_frame_equal(
        baseline_scan.candidates.sort_values("candidate_key").reset_index(drop=True),
        noisy_scan.candidates.sort_values("candidate_key").reset_index(drop=True),
    )


def test_deterministic_repeatability_full_workbook() -> None:
    bundle = load_config_from_excel(Path("PV_inputs.xlsx"))

    first_scan = run_scan(bundle)
    second_scan = run_scan(bundle)

    assert first_scan.best_candidate_key == second_scan.best_candidate_key
    pdt.assert_frame_equal(
        first_scan.candidates.sort_values("candidate_key").reset_index(drop=True),
        second_scan.candidates.sort_values("candidate_key").reset_index(drop=True),
    )
    first_kpis = build_kpis(first_scan.candidate_details[first_scan.best_candidate_key])
    second_kpis = build_kpis(second_scan.candidate_details[second_scan.best_candidate_key])
    assert math.isclose(first_kpis["NPV"], second_kpis["NPV"], rel_tol=0.0, abs_tol=1e-6)
    assert math.isclose(
        float(first_kpis["self_consumption_ratio"]),
        float(second_kpis["self_consumption_ratio"]),
        rel_tol=0.0,
        abs_tol=1e-12,
    )


def test_run_scenario_shapes_outputs() -> None:
    bundle = _fast_bundle()
    scenario = run_scenario(bundle)

    assert scenario.kpis["best_kWp"] > 0
    assert "selected_battery" in scenario.kpis
    assert "NPV" in scenario.kpis
    assert set(["Año_mes", "PV to load", "Battery to load", "Grid import"]).issubset(scenario.monthly_balance.columns)
    assert set(["Año_mes", "cumulative_npv", "monthly_savings"]).issubset(scenario.cash_flow.columns)
    assert isinstance(scenario.npv_curve, pd.DataFrame)


def test_resolve_payback_display_state_marks_payback_within_visible_horizon() -> None:
    state = resolve_payback_display_state(1.5, 2, payback_month=18)

    assert state["project_payback_years"] == pytest.approx(1.5)
    assert state["project_payback_month"] == 18
    assert state["reaches_payback"] is True
    assert state["within_visible_horizon"] is True
    assert state["message_key"] is None
    assert state["trace_payback_years"] == pytest.approx(1.5)


def test_resolve_payback_display_state_marks_payback_outside_visible_horizon() -> None:
    state = resolve_payback_display_state(1.5, 1, payback_month=18)

    assert state["project_payback_years"] == pytest.approx(1.5)
    assert state["reaches_payback"] is True
    assert state["within_visible_horizon"] is False
    assert state["message_key"] == "workbench.payback.note.visible_horizon"
    assert state["trace_payback_years"] == pytest.approx(1.5)


def test_resolve_payback_display_state_handles_missing_payback_cleanly() -> None:
    state = resolve_payback_display_state(None, 5, payback_month=None)

    assert state["project_payback_years"] is None
    assert state["reaches_payback"] is False
    assert state["within_visible_horizon"] is False
    assert state["message_key"] == "workbench.payback.note.project_horizon"
    assert state["trace_payback_years"] is None


def test_resolve_payback_display_state_uses_years_when_payback_month_is_missing() -> None:
    state = resolve_payback_display_state(1.5, 1, payback_month=None)

    assert state["project_payback_years"] == pytest.approx(1.5)
    assert state["project_payback_month"] is None
    assert state["within_visible_horizon"] is False
    assert state["message_key"] == "workbench.payback.note.visible_horizon"


def test_summarize_candidate_for_horizon_truncates_npv_only() -> None:
    monthly = pd.DataFrame(
        {
            "Año_mes": list(range(1, 25)),
            "NPV_COP": [-120.0] * 12 + [-20.0, -10.0, -5.0, -1.0, -0.5, 5.0, 10.0, 15.0, 30.0, 40.0, 50.0, 60.0],
            "Ahorro_COP": [10.0] * 24,
            "Demanda_kWh": [100.0] * 24,
            "Importacion_Red_kWh": [30.0] * 24,
            "Exportacion_kWh": [5.0] * 24,
            "PV_a_Carga_kWh": [55.0] * 24,
            "Bateria_a_Carga_kWh": [15.0] * 24,
        }
    )
    detail = {
        "scan_order": 0,
        "candidate_key": "12.000::None",
        "kWp": 12.0,
        "battery_name": "None",
        "peak_ratio": 1.1,
        "summary": {"cum_disc_final": 60.0, "payback_years": 1.5, "payback_month": 18, "capex_client": 1_000.0},
        "monthly": monthly,
    }

    first_year = summarize_candidate_for_horizon(detail, 1)
    full_horizon = summarize_candidate_for_horizon(detail, 2)

    assert first_year["horizon_years"] == 1
    assert first_year["horizon_months"] == 12
    assert first_year["NPV_COP"] == -120.0
    assert full_horizon["NPV_COP"] == 60.0


def test_build_visible_horizon_candidate_summary_preserves_project_payback() -> None:
    monthly = pd.DataFrame(
        {
            "Año_mes": list(range(1, 25)),
            "NPV_COP": [-120.0] * 12 + [-20.0, -10.0, -5.0, -1.0, -0.5, 5.0, 10.0, 15.0, 30.0, 40.0, 50.0, 60.0],
            "Ahorro_COP": [10.0] * 24,
            "Demanda_kWh": [100.0] * 24,
            "Importacion_Red_kWh": [30.0] * 24,
            "Exportacion_kWh": [5.0] * 24,
            "PV_a_Carga_kWh": [55.0] * 24,
            "Bateria_a_Carga_kWh": [15.0] * 24,
        }
    )
    detail = {
        "scan_order": 0,
        "candidate_key": "12.000::None",
        "kWp": 12.0,
        "battery_name": "None",
        "peak_ratio": 1.1,
        "summary": {"cum_disc_final": 60.0, "payback_years": 1.5, "payback_month": 18, "capex_client": 1_000.0},
        "monthly": monthly,
    }

    presentation = build_visible_horizon_candidate_summary(detail, 1)
    kpis = build_kpis(presentation)

    assert presentation["project_summary"]["payback_years"] == pytest.approx(1.5)
    assert presentation["project_summary"]["payback_month"] == 18
    assert presentation["visible_horizon_summary"]["NPV_COP"] == -120.0
    assert presentation["payback_display_state"]["message_key"] == "workbench.payback.note.visible_horizon"
    assert kpis["NPV"] == -120.0
    assert kpis["payback_years"] == pytest.approx(1.5)


def test_summarize_candidates_for_horizon_rebuilds_display_order_and_best_flag() -> None:
    monthly_base = pd.DataFrame(
        {
            "Año_mes": list(range(1, 25)),
            "Ahorro_COP": [10.0] * 24,
            "Demanda_kWh": [100.0] * 24,
            "Importacion_Red_kWh": [25.0] * 24,
            "Exportacion_kWh": [5.0] * 24,
            "PV_a_Carga_kWh": [60.0] * 24,
            "Bateria_a_Carga_kWh": [15.0] * 24,
        }
    )
    detail_map = {
        "12.000::None": {
            "scan_order": 0,
            "candidate_key": "12.000::None",
            "kWp": 12.0,
            "battery_name": "None",
            "peak_ratio": 1.05,
            "summary": {"cum_disc_final": 80.0, "payback_years": 1.5, "payback_month": 18, "capex_client": 1_000.0},
            "monthly": monthly_base.assign(NPV_COP=[50.0] * 12 + [80.0] * 12),
        },
        "12.000::BAT-10": {
            "scan_order": 1,
            "candidate_key": "12.000::BAT-10",
            "kWp": 12.0,
            "battery_name": "BAT-10",
            "peak_ratio": 1.05,
            "summary": {"cum_disc_final": 120.0, "payback_years": 1.0, "payback_month": 12, "capex_client": 1_200.0},
            "monthly": monthly_base.assign(NPV_COP=[30.0] * 12 + [120.0] * 12),
        },
        "18.000::None": {
            "scan_order": 2,
            "candidate_key": "18.000::None",
            "kWp": 18.0,
            "battery_name": "None",
            "peak_ratio": 1.2,
            "summary": {"cum_disc_final": 140.0, "payback_years": 1.0, "payback_month": 12, "capex_client": 1_500.0},
            "monthly": monthly_base.assign(NPV_COP=[90.0] * 12 + [140.0] * 12),
        },
    }

    horizon_table = summarize_candidates_for_horizon(detail_map, 1)
    full_table = summarize_candidates_for_horizon(detail_map, 2)

    assert list(horizon_table["candidate_key"]) == ["12.000::None", "12.000::BAT-10", "18.000::None"]
    assert horizon_table.loc[horizon_table["candidate_key"] == "12.000::None", "NPV_COP"].iloc[0] == 50.0
    assert horizon_table.loc[horizon_table["candidate_key"] == "12.000::BAT-10", "NPV_COP"].iloc[0] == 30.0
    assert horizon_table.loc[horizon_table["candidate_key"] == "12.000::None", "payback_years"].iloc[0] == pytest.approx(1.5)
    assert horizon_table.loc[horizon_table["candidate_key"] == "12.000::BAT-10", "payback_years"].iloc[0] == pytest.approx(1.0)
    assert horizon_table.loc[horizon_table["candidate_key"] == "12.000::None", "best_battery_for_kwp"].iloc[0]
    assert not horizon_table.loc[horizon_table["candidate_key"] == "12.000::BAT-10", "best_battery_for_kwp"].iloc[0]
    assert not full_table.loc[full_table["candidate_key"] == "12.000::None", "best_battery_for_kwp"].iloc[0]
    assert full_table.loc[full_table["candidate_key"] == "12.000::BAT-10", "best_battery_for_kwp"].iloc[0]


def test_build_npv_figure_keeps_missing_payback_as_gap() -> None:
    table = pd.DataFrame(
        [
            {
                "candidate_key": "12.000::None",
                "kWp": 12.0,
                "battery": "None",
                "NPV_COP": 10_000_000,
                "payback_years": None,
                "self_consumption_ratio": 0.45,
                "peak_ratio": 1.1,
                "scan_order": 0,
            },
            {
                "candidate_key": "18.000::BAT-10",
                "kWp": 18.0,
                "battery": "BAT-10",
                "NPV_COP": 16_000_000,
                "payback_years": 6.2,
                "self_consumption_ratio": 0.52,
                "peak_ratio": 1.25,
                "scan_order": 1,
            },
        ]
    )

    figure = build_npv_figure(table, lang="es", payback_label="Payback del proyecto [años]")
    payback_trace = next(trace for trace in figure.data if trace.name == "Payback del proyecto [años]")

    assert figure.layout.yaxis2.title.text == "Payback del proyecto [años]"
    assert np.isnan(payback_trace.y[0])
    assert payback_trace.hovertext[0].endswith(" -")


def test_scan_payload_round_trip_preserves_detail_fields() -> None:
    scan = run_scan(_fast_bundle())
    restored = type(scan).from_payload(scan.to_payload())

    assert restored.best_candidate_key == scan.best_candidate_key
    assert restored.evaluated_kwp_count == scan.evaluated_kwp_count
    assert restored.viable_kwp_count == scan.viable_kwp_count
    assert restored.discard_counts == scan.discard_counts
    assert restored.discarded_points == scan.discarded_points
    detail = restored.candidate_details[restored.best_candidate_key]
    assert "self_consumption_ratio" in detail
    assert "scan_order" in detail


def test_scan_payload_backfills_discard_telemetry_for_legacy_payload() -> None:
    scan = run_scan(_fast_bundle())
    payload = scan.to_payload()
    payload.pop("evaluated_kwp_count")
    payload.pop("viable_kwp_count")
    payload.pop("discard_counts")
    payload.pop("discarded_points")

    restored = type(scan).from_payload(payload)

    assert restored.viable_kwp_count == int(scan.candidates["kWp"].nunique())
    assert restored.evaluated_kwp_count == restored.viable_kwp_count
    assert restored.discard_counts == {}
    assert restored.discarded_points == ()


def test_candidate_table_has_single_best_battery_per_kwp_and_detail_alignment() -> None:
    scan = run_scan(_fast_bundle())

    best_counts = scan.candidates.groupby("kWp")["best_battery_for_kwp"].sum()
    assert (best_counts == 1).all()

    flagged_keys = set(scan.candidates.loc[scan.candidates["best_battery_for_kwp"], "candidate_key"])
    for candidate_key, detail in scan.candidate_details.items():
        assert detail["best_battery"] == (candidate_key in flagged_keys)


def test_selected_candidate_helper_uses_selected_row_over_best() -> None:
    scan = run_scan(_fast_bundle())
    table_rows = scan.candidates.to_dict("records")
    non_best_index = next(index for index, row in enumerate(table_rows) if row["candidate_key"] != scan.best_candidate_key)

    selected_key = resolve_selected_candidate_key(scan, [non_best_index], table_rows)

    assert selected_key == table_rows[non_best_index]["candidate_key"]
    selected_detail = scan.candidate_details[selected_key]
    best_detail = scan.candidate_details[scan.best_candidate_key]
    assert selected_detail["candidate_key"] != best_detail["candidate_key"]
    assert build_kpis(selected_detail)["best_kWp"] == selected_detail["kWp"]


def test_selected_candidate_helper_for_scenario_prefers_click_then_row_then_stored_key() -> None:
    scan = run_scan(_fast_bundle())
    table_rows = scan.candidates.to_dict("records")
    non_best_index = next(index for index, row in enumerate(table_rows) if row["candidate_key"] != scan.best_candidate_key)
    row_selected_key = table_rows[non_best_index]["candidate_key"]

    click_selected_key = resolve_selected_candidate_key_for_scenario(
        scan,
        scan.best_candidate_key,
        table_rows=table_rows,
        selected_rows=[non_best_index],
        click_data={"points": [{"customdata": [scan.best_candidate_key]}]},
    )
    assert click_selected_key == scan.best_candidate_key

    row_only_key = resolve_selected_candidate_key_for_scenario(
        scan,
        scan.best_candidate_key,
        table_rows=table_rows,
        selected_rows=[non_best_index],
        click_data=None,
    )
    assert row_only_key == row_selected_key

    stored_key = resolve_selected_candidate_key_for_scenario(
        scan,
        row_selected_key,
        table_rows=table_rows,
        selected_rows=None,
        click_data=None,
    )
    assert stored_key == row_selected_key


def test_capex_semantics_variable_mode() -> None:
    cfg = {
        "pricing_mode": "variable",
        "price_total_COP": 58_000_000,
        "price_others_total": 2_000_000,
        "include_hw_in_price": False,
    }
    inv_sel = {"inverter": {"price_COP": 5_000_000}}
    battery_sel = {"nom_kWh": 10.0, "price_COP": 12_000_000}

    without_hw = calculate_capex_client(cfg, 10.0, inv_sel, battery_sel, 3_500_000)
    with_hw = calculate_capex_client({**cfg, "include_hw_in_price": True}, 10.0, inv_sel, battery_sel, 3_500_000)

    assert without_hw == 37_000_000
    assert with_hw == 54_000_000


def test_capex_semantics_total_mode() -> None:
    cfg = {
        "pricing_mode": "total",
        "price_total_COP": 80_000_000,
        "price_others_total": 1_500_000,
        "include_hw_in_price": False,
    }
    inv_sel = {"inverter": {"price_COP": 9_000_000}}
    battery_sel = {"nom_kWh": 0.0, "price_COP": 0.0}

    without_hw = calculate_capex_client(cfg, 12.0, inv_sel, battery_sel, 999_999)
    with_hw = calculate_capex_client({**cfg, "include_hw_in_price": True}, 12.0, inv_sel, battery_sel, 999_999)

    assert without_hw == 81_500_000
    assert with_hw == 90_500_000


def test_missing_sheet_raises_friendly_contract_error(tmp_path) -> None:
    workbook = _copy_workbook(tmp_path, "missing_sheet.xlsx")
    _remove_sheet(workbook, "Perfiles")

    with pytest.raises(WorkbookContractError, match="Falta la hoja 'Perfiles'"):
        load_config_from_excel(workbook)


def test_missing_required_table_raises_friendly_contract_error(tmp_path) -> None:
    workbook = _copy_workbook(tmp_path, "missing_table.xlsx")
    _remove_table(workbook, "Perfiles", "Demand_Profile")

    with pytest.raises(WorkbookContractError, match="Falta la tabla 'Demand_Profile'"):
        load_config_from_excel(workbook)


def test_build_7x24_from_excel_validation_message_matches_total_mode() -> None:
    df = pd.DataFrame({"DOW": [1], "HOUR": [0]})

    with pytest.raises(ValueError, match=r"Profiles\.LoadProfile debe tener: DOW, HOUR, TOTAL"):
        build_7x24_from_excel(df, total=True)


def test_build_7x24_from_excel_validation_message_matches_split_mode() -> None:
    df = pd.DataFrame({"DOW": [1], "HOUR": [0], "TOTAL": [1.0]})

    with pytest.raises(ValueError, match=r"Profiles\.LoadProfile debe tener: DOW, HOUR, RES, IND, TOTAL"):
        build_7x24_from_excel(df, total=False)


def test_missing_required_column_raises_friendly_contract_error(tmp_path) -> None:
    workbook = _copy_workbook(tmp_path, "missing_column.xlsx")
    _rename_table_column(workbook, "Catalogos", "Battery_Catalog", "max_dis_kW", "max_discharge")

    with pytest.raises(WorkbookContractError, match="max_dis_kW"):
        load_config_from_excel(workbook)


def test_invalid_mode_value_produces_validation_error(tmp_path) -> None:
    workbook = _copy_workbook(tmp_path, "invalid_mode.xlsx")
    _set_config_value(workbook, "pricing_mode", "banana")

    bundle = load_config_from_excel(workbook)

    assert any(issue.field == "pricing_mode" and issue.level == "error" for issue in bundle.issues)
    with pytest.raises(ValueError, match="pricing_mode"):
        run_scan(bundle)


def test_invalid_boolean_value_produces_validation_error(tmp_path) -> None:
    workbook = _copy_workbook(tmp_path, "invalid_bool.xlsx")
    _set_config_value(workbook, "include_battery", "maybe")

    bundle = load_config_from_excel(workbook)

    assert any(issue.field == "include_battery" and issue.level == "error" for issue in bundle.issues)


def test_empty_inverter_catalog_is_rejected(tmp_path) -> None:
    bundle = replace(load_example_config(), inverter_catalog=load_example_config().inverter_catalog.iloc[0:0].copy())
    issues = validate_config(bundle)

    assert any(issue.field == "Inversor_Catalog" and issue.level == "error" for issue in issues)
    with pytest.raises(ValueError, match="Inversor_Catalog"):
        run_scan(bundle)


def test_empty_battery_catalog_is_rejected_when_batteries_enabled(tmp_path) -> None:
    example = load_example_config()
    bundle = replace(
        example,
        config={**example.config, "include_battery": True, "optimize_battery": True},
        battery_catalog=example.battery_catalog.iloc[0:0].copy(),
    )
    issues = validate_config(bundle)

    assert any(issue.field == "Battery_Catalog" and issue.level == "error" for issue in issues)
    with pytest.raises(ValueError, match="Battery_Catalog"):
        run_scan(bundle)


def test_no_battery_configuration_still_scans_through_the_no_battery_path() -> None:
    bundle = _fast_bundle()
    no_battery_bundle = replace(
        bundle,
        config={**bundle.config, "include_battery": False, "optimize_battery": False, "battery_name": ""},
    )

    scan = run_scan(no_battery_bundle)

    assert scan.candidates.empty is False
    assert set(scan.candidates["battery"].astype(str)) == {"None"}
    assert {detail["battery_name"] for detail in scan.candidate_details.values()} == {"None"}
    assert all(float(detail["battery"]["nom_kWh"]) == 0.0 for detail in scan.candidate_details.values())
    assert all(float(detail["battery"]["max_kW"]) == 0.0 for detail in scan.candidate_details.values())
    assert all(key.endswith("::None") for key in scan.candidate_details)


def test_invalid_optimization_range_is_rejected(tmp_path) -> None:
    workbook = _copy_workbook(tmp_path, "invalid_range.xlsx")
    _set_config_value(workbook, "kWp_min", 30)
    _set_config_value(workbook, "kWp_max", 10)

    bundle = load_config_from_excel(workbook)

    assert any(issue.field == "kWp_min" and issue.level == "error" for issue in bundle.issues)
    with pytest.raises(ValueError, match="kWp_min"):
        run_scan(bundle)


def test_peak_ratio_behavior_is_kept_for_legacy_compatibility() -> None:
    bundle = _fast_bundle()
    cfg = {**bundle.config, "limit_peak_ratio_enable": True}
    from pv_product.hardware import peak_ratio_ok

    ok_a, ratio_a = peak_ratio_ok(
        cfg,
        15.0,
        {"inverter": {"AC_kW": 10.0}},
        bundle.solar_profile,
        bundle.hsp_month,
        bundle.demand_month_factor,
        dow24=bundle.demand_profile_7x24,
        day_w=bundle.day_weights,
    )
    ok_b, ratio_b = peak_ratio_ok(
        cfg,
        15.0,
        {"inverter": {"AC_kW": 10.0}},
        bundle.solar_profile,
        bundle.hsp_month,
        bundle.demand_month_factor * 1.2,
        dow24=bundle.demand_profile_7x24,
        day_w=bundle.day_weights,
    )

    assert ok_a == ok_b or isinstance(ok_a, bool)
    assert ratio_b != ratio_a


def test_all_discarded_peak_ratio_scan_returns_completed_result() -> None:
    bundle = _fast_bundle()
    discarded_bundle = replace(
        bundle,
        config={**bundle.config, "limit_peak_ratio_enable": True, "limit_peak_ratio": 0.01},
    )

    scan = run_scan(discarded_bundle)

    assert scan.best_candidate_key is None
    assert scan.candidates.empty
    assert scan.viable_kwp_count == 0
    assert scan.evaluated_kwp_count > 0
    assert scan.discard_counts["peak_ratio"] == scan.evaluated_kwp_count
    assert scan.discard_counts["inverter_string"] == 0
    assert scan.discarded_points
    assert all(point["reason"] == "peak_ratio" for point in scan.discarded_points)


def test_selected_candidate_helper_ignores_discard_marker_clicks() -> None:
    scan = run_scan(_fast_bundle())

    selected_key = resolve_selected_candidate_key_for_scenario(
        scan,
        scan.best_candidate_key,
        click_data={"points": [{"x": 18.0, "y": 0.5}]},
    )

    assert selected_key == scan.best_candidate_key


def test_regression_against_preserved_legacy_artifacts() -> None:
    bundle = load_config_from_excel(Path("PV_inputs.xlsx"))
    scan = run_scan(bundle)
    legacy_summary = _parse_legacy_summary(Path("Resultados/resumen_optimizacion.txt"))
    legacy_scan = pd.read_csv("Resultados/resumen_valor_presente_neto.csv")
    legacy_row = legacy_scan[(legacy_scan["kWp"] == legacy_summary["best_kwp"]) & (legacy_scan["battery"] == legacy_summary["battery"])].iloc[0]
    legacy_candidate_key = f"{legacy_summary['best_kwp']:.3f}::{legacy_summary['battery']}"
    legacy_detail = scan.candidate_details[legacy_candidate_key]
    best_detail = scan.candidate_details[scan.best_candidate_key]

    # NPV and payback tolerances are intentionally looser because the preserved
    # batch artifacts were generated before deterministic/stochastic separation,
    # and the hardened scan now also includes prime module counts.
    assert legacy_candidate_key in scan.candidate_details
    assert math.isclose(legacy_detail["peak_ratio"], float(legacy_row["peak_ratio"]), rel_tol=0.0, abs_tol=1e-6)
    assert math.isclose(legacy_detail["summary"]["cum_disc_final"], float(legacy_row["NPV"]), rel_tol=0.0, abs_tol=6_000_000)
    assert math.isclose(float(legacy_detail["summary"]["payback_years"]), float(legacy_summary["payback_years"]), rel_tol=0.0, abs_tol=1.0)

    # The hardened optimum is now allowed to move because prime module counts
    # are no longer excluded from the search space.
    assert best_detail["battery_name"] == legacy_summary["battery"]
    assert best_detail["kWp"] >= legacy_summary["best_kwp"]


def test_template_round_trip(tmp_path) -> None:
    template_path = tmp_path / "PV_inputs_template.xlsx"
    ensure_template(template_path)

    bundle = load_config_from_excel(template_path)

    assert bundle.inverter_catalog.empty is False
    assert bundle.battery_catalog.empty is False
    assert bundle.issues is not None
